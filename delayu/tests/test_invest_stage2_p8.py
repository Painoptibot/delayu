from __future__ import annotations

from datetime import datetime, timezone as datetime_timezone
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from delayu.models import ModuleCatalog, Organization, Role, RoleModulePermission, Subsystem, SubsystemMembership, SubsystemModule
from delayu.models_invest import InvestAutomationConfig, InvestProject, InvestRoadmapItem
from delayu.services.invest_dashboard import build_dashboard
from delayu.services.invest_flags import ensure_automation_config
from delayu.services.invest_roles import perm_for_role

User = get_user_model()


@pytest.fixture
def p8_ctx(db):
    sub = Subsystem.objects.create(
        code="inv-p8", name="Invest P8", industry_template="invest", status=Subsystem.Status.ACTIVE
    )
    module = ModuleCatalog.objects.create(code="M22", name="Инвестпроекты")
    SubsystemModule.objects.create(subsystem=sub, module=module, enabled=True)
    org = Organization.objects.create(subsystem=sub, code="mo1", name="МО-1")
    role = Role.objects.create(subsystem=sub, code="invest_admin", name="Admin")
    RoleModulePermission.objects.create(role=role, module=module, **perm_for_role(role.code, "M22"))
    user = User.objects.create_user("p8-admin", password="x", is_superuser=True)
    SubsystemMembership.objects.create(user=user, subsystem=sub, organization=org, role=role, is_default=True)
    cfg = ensure_automation_config(sub)
    return {"sub": sub, "org": org, "role": role, "user": user, "cfg": cfg}


def _project(ctx, *, code, name=None, funnel=InvestProject.Funnel.ATTRACTION, stage="lead", industry="", amount=None, jobs=None):
    return InvestProject.objects.create(
        subsystem=ctx["sub"],
        organization=ctx["org"],
        code=code,
        name=name or code,
        investor_name="ООО Инвестор",
        funnel=funnel,
        stage=stage,
        industry=industry,
        investment_amount=amount,
        jobs_count=jobs,
    )


@pytest.mark.django_db
def test_dashboard_includes_industry_investment_and_period_compare(p8_ctx):
    now = timezone.now()
    project = _project(p8_ctx, code="P8-1", industry="АПК", amount="10.50", jobs=12)
    old_project = _project(p8_ctx, code="P8-2", industry="Промышленность", amount="7.00", jobs=5)
    InvestProject.objects.filter(pk=project.pk).update(created_at=now - timezone.timedelta(days=2))
    InvestProject.objects.filter(pk=old_project.pk).update(created_at=now - timezone.timedelta(days=12))

    dashboard = build_dashboard(p8_ctx["sub"], period="week", now=now)

    assert dashboard["investment_total"] == pytest.approx(17.5)
    assert dashboard["jobs_total"] == 17
    assert dashboard["industry_metrics"][0]["industry"] == "АПК"
    assert dashboard["industry_metrics"][0]["projects"] == 1
    assert dashboard["period_compare"]["current_count"] == 1
    assert dashboard["period_compare"]["previous_count"] == 1


@pytest.mark.django_db
def test_dashboard_export_downloads_xlsx_with_pack_sections(client, p8_ctx):
    project = _project(p8_ctx, code="P8-XLSX", stage="lead", industry="АПК", amount="9.00", jobs=3)
    InvestRoadmapItem.objects.create(
        project=project,
        code="r1",
        title="Просроченный шаг",
        status=InvestRoadmapItem.Status.OVERDUE,
        due_at=timezone.now() - timezone.timedelta(days=1),
    )
    client.force_login(p8_ctx["user"])

    resp = client.get(reverse("invest-dashboard-export"))

    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    wb = load_workbook(BytesIO(resp.content))
    assert {"Funnel", "Overdue", "Bottlenecks", "Industry"}.issubset(set(wb.sheetnames))
    assert wb["Funnel"]["A2"].value == "Привлечение"
    assert wb["Overdue"]["B2"].value == 1
    assert wb["Industry"]["A2"].value == "АПК"


@pytest.mark.django_db
def test_kanban_groups_attraction_and_support_cards(client, p8_ctx):
    attraction = _project(p8_ctx, code="KAN-1", name="Attract", stage="lead")
    support = _project(
        p8_ctx,
        code="KAN-2",
        name="Support",
        funnel=InvestProject.Funnel.SUPPORT,
        stage="accepted",
    )
    client.force_login(p8_ctx["user"])

    resp = client.get(reverse("invest-kanban"))

    assert resp.status_code == 200
    assert attraction in resp.context["kanban"]["attraction"]["lead"]["projects"]
    assert support in resp.context["kanban"]["support"]["accepted"]["projects"]
    assert reverse("invest-project-detail", args=[attraction.pk]).encode() in resp.content


@pytest.mark.django_db
def test_dashboard_period_query_uses_from_to_created_range(client, p8_ctx):
    inside = _project(p8_ctx, code="PER-IN")
    outside = _project(p8_ctx, code="PER-OUT")
    InvestProject.objects.filter(pk=inside.pk).update(
        created_at=datetime(2026, 7, 10, 12, tzinfo=datetime_timezone.utc)
    )
    InvestProject.objects.filter(pk=outside.pk).update(
        created_at=datetime(2026, 5, 1, 12, tzinfo=datetime_timezone.utc)
    )
    client.force_login(p8_ctx["user"])

    resp = client.get(reverse("invest-dashboard"), {"from": "2026-07-01", "to": "2026-07-31"})

    assert resp.status_code == 200
    assert resp.context["dashboard"]["period_compare"]["current_count"] == 1
    assert resp.context["dashboard"]["period_compare"]["previous_count"] == 0


@pytest.mark.django_db
def test_escalation_rules_page_saves_config_options(client, p8_ctx):
    client.force_login(p8_ctx["user"])

    resp = client.post(
        reverse("invest-escalation-rules"),
        {"due_days": "4", "levels": "МО\nДепартамент\nГубернатор"},
    )

    assert resp.status_code == 302
    cfg = InvestAutomationConfig.objects.get(subsystem=p8_ctx["sub"])
    assert cfg.options["escalation_rules"] == {
        "due_days": 4,
        "levels": ["МО", "Департамент", "Губернатор"],
    }
