"""Views for the invest subsystem."""
from io import BytesIO
import json
import os
import re

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import AccessMixin
from django.core.exceptions import PermissionDenied
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import Q
from django.http import Http404, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, DetailView, ListView, TemplateView, UpdateView
from django.views.generic.base import ContextMixin
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from delayu.forms_invest import (
    InvestProjectForm,
    InvestProjectWizardStep1Form,
    InvestProjectWizardStep2Form,
    InvestProjectWizardStep3Form,
    InvestSiteForm,
    InvestSupportTrackItemForm,
    OKVED_INDUSTRIES,
)
from delayu.mixins import ModulePermissionMixin
from delayu.models import AuditLog, DocumentFile, Subsystem, UserProfile
from delayu.models_invest import (
    InvestExternalTask,
    InvestExtract,
    InvestFgistpDocument,
    InvestFgistpRecord,
    InvestHandoff,
    InvestImportBatch,
    InvestImportRow,
    InvestInvestor,
    InvestPackageItem,
    InvestProject,
    InvestProjectComment,
    InvestProjectSite,
    InvestRoadmapItem,
    InvestSite,
    InvestSmevRequest,
    InvestStopFactor,
)
from delayu.services.access import get_membership_or_403, user_can
from delayu.services import audit
from delayu.services.invest_booking import InvestBookingError, book_site, expire_overdue_bookings, select_site
from delayu.services.invest_dashboard import build_cockpit, build_dashboard, project_sla_risk
from delayu.services.invest_dedup import ignore_duplicate_pair, inn_is_valid, suspected_duplicate_pairs
from delayu.services.invest_escalation import refresh_invest_sla
from delayu.services.invest_opendata import (
    run_investor_verification,
    run_project_verification,
    run_site_verification,
)
from delayu.services.invest_opendata.orchestrator import InvestOpenDataError
from delayu.services.invest_extract_mnp import (
    intersections_from_extract,
    refresh_extract_mnp_intersections,
)
from delayu.services.invest_extracts import (
    InvestExtractError,
    attach_extract_to_package,
    ensure_extract_for_site,
    extract_geometry_for_map,
    extracts_for_inbox,
    generate_mock_contour,
    import_extract_geometry,
    latest_map_extract,
    mark_extract_received,
    verify_extract,
)
from delayu.services.invest_fgistp import (
    InvestFgistpError,
    attach_fgistp_document,
    attach_fgistp_to_package,
    ensure_fgistp_for_site,
    fgistp_for_inbox,
    fgistp_geometry_for_map,
    generate_mock_zones,
    import_fgistp_geometry,
    latest_map_fgistp,
    mark_fgistp_received,
    search_fgistp_documents,
    verify_fgistp,
)
from delayu.services.invest_mnp import InvestMnpError, fetch_mnp_wfs, fetch_mnp_wms, mnp_map_config
from delayu.services.invest_mnp_store import (
    query_features_geojson,
    read_tile_bytes,
    render_viewport_png,
    store_status,
)
from delayu.services.invest_mnp_styles import map_style_config
from delayu.services.invest_bitrix import InvestBitrixError, push_project_to_bitrix, resolve_bitrix_stage_conflict
from delayu.services.invest_gates import can_push_to_bitrix, compute_completeness, gate_blockers
from delayu.services.invest_handoff import (
    InvestHandoffError,
    RETURN_REASON_TEMPLATES,
    accept_handoff,
    request_handoff,
    resolve_return_comment,
    return_handoff,
)
from delayu.services.invest_import import apply_row, parse_mo_file, skip_row
from delayu.services.invest_flags import ensure_automation_config, flag_enabled
from delayu.services.invest_package import ensure_package, set_item_status
from delayu.services.invest_scope import projects_for_membership, sites_for_membership
from delayu.services.invest_pipeline import auto_smev_enrich_site
from delayu.services.invest_smev import (
    InvestSmevError,
    apply_smev_response,
    batch_smev_requests,
    build_smev_report,
    emulate_gateway_response,
    render_smev_protocol_pdf,
    request_contour_check,
    request_smev_fill,
    user_can_apply_live,
)
from delayu.services.odysseus_invest import get_invest_odysseus_open_url, prepare_odysseus_open
from delayu.services.scope import is_platform_admin
from delayu.services.yandex_maps import YandexGeocodeError, geocode_address


PROJECT_WIZARD_SESSION_KEY = "invest_project_wizard"

ROLE_HOMES = {
    "invest_agency": {
        "title": "Агентство развития",
        "blurb": "Фокус на привлечении инвесторов, квалификации лидов и подготовке передачи в сопровождение.",
        "widgets": [
            {"label": "Новые лиды", "metric": "attraction"},
            {"label": "Передачи", "url_name": "invest-handoffs"},
            {"label": "Площадки", "url_name": "invest-sites"},
        ],
    },
    "invest_dept": {
        "title": "Департамент инвестиций",
        "blurb": "Контроль воронки, SLA, пакетов передачи и узких мест по муниципалитетам.",
        "widgets": [
            {"label": "Дашборд", "url_name": "invest-dashboard"},
            {"label": "Канбан", "url_name": "invest-kanban"},
            {"label": "Просрочки", "url_name": "invest-inbox"},
        ],
    },
    "invest_mo": {
        "title": "Муниципальный кабинет",
        "blurb": "Работа с проектами и площадками своего МО, актуализация сведений и ответы на запросы.",
        "widgets": [
            {"label": "Мои проекты", "metric": "projects_total"},
            {"label": "Площадки МО", "url_name": "invest-sites"},
            {"label": "Сегодня", "url_name": "invest-inbox"},
        ],
    },
    "invest_admin": {
        "title": "Администрирование инвестконтура",
        "blurb": "Настройка интеграций, автоматизации и сквозного контроля инвестконтура.",
        "widgets": [
            {"label": "Автоматизация", "url_name": "invest-automation"},
            {"label": "Интеграции", "url_name": "platform-integrations"},
            {"label": "Дашборд", "url_name": "invest-dashboard"},
        ],
    },
}


def _role_home_for_membership(membership):
    role_code = membership.role.code
    home = dict(ROLE_HOMES.get(role_code) or ROLE_HOMES["invest_agency"])
    role_homes = (ensure_automation_config(membership.subsystem).options or {}).get("role_homes") or {}
    override = role_homes.get(role_code) or {}
    for key in ("title", "blurb"):
        value = (override.get(key) or "").strip()
        if value:
            home[key] = value
    home["role_code"] = role_code
    return home


INVEST_ONBOARDING_STEPS = [
    {"id": "hub", "title": "Откройте обзор инвестконтура", "url_name": "invest-hub"},
    {"id": "inbox", "title": "Проверьте inbox Сегодня", "url_name": "invest-inbox"},
    {"id": "project_list", "title": "Откройте реестр проектов", "url_name": "invest-projects"},
    {"id": "project_detail", "title": "Изучите карточку проекта", "url_name": "invest-projects"},
    {"id": "dashboard", "title": "Посмотрите дашборд", "url_name": "invest-dashboard"},
    {"id": "kanban", "title": "Проверьте канбан", "url_name": "invest-kanban"},
    {"id": "sites", "title": "Откройте каталог площадок", "url_name": "invest-sites"},
    {"id": "handoffs", "title": "Проверьте передачи", "url_name": "invest-handoffs"},
    {"id": "automation", "title": "Проверьте автоматизацию", "url_name": "invest-automation"},
    {"id": "comments", "title": "Оставьте рабочий комментарий", "url_name": "invest-projects"},
]


def _invest_onboarding_state(user):
    profile, _ = UserProfile.objects.get_or_create(user=user, defaults={"is_platform_admin": False})
    state = dict(profile.onboarding_state or {})
    invest = dict(state.get("invest") or {})
    invest.setdefault("completed", [])
    return profile, state, invest


class InvestSubsystemMixin(AccessMixin):
    module_code = "M22"
    page_title = "Инвестконтур"

    def get_membership(self):
        if not hasattr(self, "_invest_membership"):
            self._invest_membership = get_membership_or_403(self.request)
        return self._invest_membership

    def get_subsystem(self):
        return self.get_membership().subsystem

    def dispatch(self, request, *args, **kwargs):
        # До ModulePermissionMixin: иначе AnonymousUser уходит в filter(user=…).
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        membership = get_membership_or_403(request)
        if (
            membership.subsystem.industry_template != "invest"
            or membership.subsystem.status != Subsystem.Status.ACTIVE
        ):
            messages.error(request, "Раздел доступен только в активном инвестконтуре. Переключите контур.")
            return redirect("platform-home")
        self._invest_membership = membership
        return super().dispatch(request, *args, **kwargs)


class InvestForbiddenResponseMixin:
    def dispatch(self, request, *args, **kwargs):
        if request.method in {"POST", "PUT", "PATCH", "DELETE"} and not user_can(
            request.user, self.module_code, self.required_action
        ):
            return HttpResponseForbidden(f"Нет доступа к {self.module_code}")
        return super().dispatch(request, *args, **kwargs)


class InvestImportRoleMixin:
    allowed_role_codes = {"invest_mo", "invest_dept", "invest_admin"}

    def dispatch(self, request, *args, **kwargs):
        membership = get_membership_or_403(request)
        if membership.role.code not in self.allowed_role_codes:
            return HttpResponseForbidden("Импорт доступен только ролям МО, департамента и администратора")
        return super().dispatch(request, *args, **kwargs)


def _import_batches_for_membership(membership):
    qs = InvestImportBatch.objects.filter(subsystem=membership.subsystem).select_related("organization")
    if membership.role.code == "invest_mo":
        qs = qs.filter(organization=membership.organization)
    return qs


def _membership_has_role(membership, allowed_role_codes):
    return membership.role.code in allowed_role_codes


def _forbidden_with_message(request, message):
    messages.error(request, message)
    return HttpResponseForbidden(message)


def _with_odysseus_cta(request, ctx, *, membership, project=None, site=None):
    ctx["odysseus_cta_url"] = get_invest_odysseus_open_url(
        request,
        membership=membership,
        project=project,
        site=site,
    )
    return ctx


def _stage_to_funnel() -> dict[str, str]:
    result = {}
    for funnel, transitions in InvestProjectForm.STAGE_TRANSITIONS.items():
        for stage, allowed in transitions.items():
            result[stage] = funnel
            for next_stage in allowed:
                result[next_stage] = funnel
    return result


def _can_bulk_update_stage(user, membership) -> bool:
    return (
        getattr(user, "is_superuser", False)
        or is_platform_admin(user)
        or membership.role.code in {"invest_admin", "invest_dept"}
    )


def _can_admin_override(user, membership) -> bool:
    return (
        getattr(user, "is_superuser", False)
        or is_platform_admin(user)
        or membership.role.code in {"invest_admin", "invest_dept"}
    )


def _completeness_actions(project, blockers):
    labels = {
        "name": "Добавьте наименование проекта",
        "investor_name": "Добавьте инвестора",
        "organization_id": "Укажите МО / территорию",
        "industry": "Добавьте отрасль",
        "package_incomplete": "Заполните обязательные пункты пакета",
        "mo_pending": "Закройте открытые задачи МО",
        "stop_factor": "Устраните открытые стоп-факторы",
    }
    actions = [labels.get(blocker, blocker) for blocker in blockers]
    if not project.contact_person:
        actions.append("Добавьте контактное лицо")
    if not project.contact_phone:
        actions.append("Добавьте телефон контакта")
    if not project.description:
        actions.append("Добавьте описание проекта")
    if project.investment_amount is None:
        actions.append("Укажите объём инвестиций")
    return list(dict.fromkeys(actions))


def _open_stop_factors(project):
    return project.stop_factors.filter(
        status__in=(InvestStopFactor.Status.OPEN, InvestStopFactor.Status.BLOCKING),
    )


def _passport_pdf_fonts():
    regular_name = "Helvetica"
    bold_name = "Helvetica-Bold"
    candidates = [
        (r"C:\Windows\Fonts\arial.ttf", r"C:\Windows\Fonts\arialbd.ttf"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    ]
    for regular_path, bold_path in candidates:
        if os.path.exists(regular_path) and os.path.exists(bold_path):
            if "DelayuPassport" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("DelayuPassport", regular_path))
            if "DelayuPassportBold" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("DelayuPassportBold", bold_path))
            regular_name = "DelayuPassport"
            bold_name = "DelayuPassportBold"
            break
    return regular_name, bold_name


def _opendata_snapshot_from_entity(entity) -> dict | None:
    if entity is None:
        return None
    if hasattr(entity, "extras"):
        data = (entity.extras or {}).get("opendata")
        if isinstance(data, dict):
            return data
    if hasattr(entity, "external_ids"):
        data = (entity.external_ids or {}).get("opendata")
        if isinstance(data, dict):
            return data
    return None


def _with_opendata_context(ctx, *, entity, verify_url: str, can_verify: bool):
    ctx["opendata_snapshot"] = _opendata_snapshot_from_entity(entity)
    ctx["opendata_verify_url"] = verify_url
    ctx["opendata_can_verify"] = can_verify
    return ctx


def _coordinates_for_site(site):
    if site.latitude is not None and site.longitude is not None:
        return str(site.latitude), str(site.longitude)
    external = site.external_ids or {}
    coordinates = external.get("coordinates") or external.get("coords")
    if isinstance(coordinates, dict):
        lat = coordinates.get("lat") or coordinates.get("latitude")
        lon = coordinates.get("lon") or coordinates.get("lng") or coordinates.get("longitude")
        if lat and lon:
            return str(lat), str(lon)
    if isinstance(coordinates, (list, tuple)) and len(coordinates) >= 2:
        return str(coordinates[0]), str(coordinates[1])
    if isinstance(coordinates, str) and "," in coordinates:
        lat, lon = coordinates.split(",", 1)
        return lat.strip(), lon.strip()
    return None, None


def _restriction_zone_name(zone):
    if isinstance(zone, dict):
        return str(zone.get("name") or "").strip()
    return str(zone or "").strip()


def _restriction_zone_coords(zone):
    if not isinstance(zone, dict):
        return None
    coords = zone.get("coords")
    if not isinstance(coords, list) or len(coords) < 3:
        return None
    normalized = []
    for pair in coords:
        if not isinstance(pair, (list, tuple)) or len(pair) < 2:
            return None
        try:
            normalized.append([float(pair[0]), float(pair[1])])
        except (TypeError, ValueError):
            return None
    return normalized


def _restriction_zones_for_map(site):
    zones = []
    for zone in site.restriction_zones or []:
        name = _restriction_zone_name(zone)
        if not name:
            continue
        entry = {"name": name}
        coords = _restriction_zone_coords(zone)
        if coords:
            entry["coords"] = coords
        zones.append(entry)
    return zones


def _booking_badge_for_site(site):
    info = _booking_info_for_site(site)
    return info.get("badge") or ""


def _booking_info_for_site(site):
    links = list(getattr(site, "prefetched_project_links", None) or [])
    if not links:
        links = list(site.project_links.all())
    if not links:
        return {}
    priority = {
        InvestProjectSite.Role.SELECTED: 0,
        InvestProjectSite.Role.BOOKED: 1,
        InvestProjectSite.Role.PROPOSED: 2,
        InvestProjectSite.Role.CANDIDATE: 3,
    }
    link = sorted(links, key=lambda item: priority.get(item.role, 99))[0]
    project = getattr(link, "project", None)
    return {
        "badge": InvestProjectSite.Role(link.role).label,
        "projectName": getattr(project, "name", "") or "",
        "projectCode": getattr(project, "code", "") or "",
        "projectUrl": reverse("invest-project-detail", args=[project.pk]) if project else "",
    }


def _smev_payload_summary(payload):
    if not payload:
        return "—"
    keys = ("source", "area_ha", "vri", "land_category", "right_type", "received_at")
    parts = [f"{key}: {payload[key]}" for key in keys if payload.get(key)]
    if parts:
        return "; ".join(parts)
    return "; ".join(f"{key}: {value}" for key, value in list(payload.items())[:4])


def _smev_field_diff(payload):
    diff = (payload or {}).get("field_diff") or {}
    return [
        {"field": field, "old": values.get("old") or "—", "new": values.get("new") or "—"}
        for field, values in diff.items()
        if isinstance(values, dict)
    ]


def _campaign_progress(sites):
    campaigns = {}
    for site in sites:
        refresh = (site.external_ids or {}).get("mo_refresh") or {}
        campaign_id = refresh.get("campaign_id")
        if not campaign_id:
            continue
        campaign = campaigns.setdefault(
            campaign_id,
            {
                "campaign_id": campaign_id,
                "organization": site.organization,
                "requested_at": refresh.get("requested_at", ""),
                "counts": {},
                "total": 0,
            },
        )
        status = refresh.get("status") or "queued"
        campaign["counts"][status] = campaign["counts"].get(status, 0) + 1
        campaign["total"] += 1
    return sorted(campaigns.values(), key=lambda row: row["requested_at"], reverse=True)


class InvestHubView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/hub.html"
    page_title = "Обзор инвестконтура"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        projects = projects_for_membership(membership)
        ctx["stats"] = {
            "projects_total": projects.count(),
            "attraction": projects.filter(funnel=InvestProject.Funnel.ATTRACTION).count(),
            "support": projects.filter(funnel=InvestProject.Funnel.SUPPORT).count(),
        }
        ctx["recent_projects"] = projects.select_related("organization", "owner")[:8]
        ctx["can_create_project"] = user_can(self.request.user, self.module_code, "create")
        ctx["role_home"] = _role_home_for_membership(membership)
        return _with_odysseus_cta(self.request, ctx, membership=membership)


class InvestOnboardingView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/onboarding.html"
    page_title = "Онбординг инвестконтура"
    http_method_names = ["get", "post", "head", "options"]

    def post(self, request, *args, **kwargs):
        step_id = (request.POST.get("step_id") or "").strip()
        valid_ids = {step["id"] for step in INVEST_ONBOARDING_STEPS}
        if step_id in valid_ids:
            profile, state, invest = _invest_onboarding_state(request.user)
            completed = list(invest.get("completed") or [])
            if step_id not in completed:
                completed.append(step_id)
            invest["completed"] = completed
            state["invest"] = invest
            profile.onboarding_state = state
            profile.save(update_fields=["onboarding_state", "updated_at"])
            messages.success(request, "Шаг онбординга отмечен.")
        return redirect(reverse("invest-onboarding"))

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        _, _, invest = _invest_onboarding_state(self.request.user)
        completed = set(invest.get("completed") or [])
        steps = []
        for index, step in enumerate(INVEST_ONBOARDING_STEPS, start=1):
            row = dict(step)
            row["index"] = index
            row["done"] = row["id"] in completed
            row["url"] = reverse(row["url_name"])
            steps.append(row)
        ctx["invest_onboarding_steps"] = steps
        ctx["invest_onboarding_completed"] = len(completed.intersection({step["id"] for step in steps}))
        ctx["invest_onboarding_total"] = len(steps)
        return ctx


class InvestDashboardView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/dashboard.html"
    page_title = "Дашборд руководителя"

    def get_dashboard_kwargs(self):
        return {
            "period": self.request.GET.get("period") or "",
            "date_from": parse_date(self.request.GET.get("from") or ""),
            "date_to": parse_date(self.request.GET.get("to") or ""),
        }

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["dashboard"] = build_dashboard(self.get_subsystem(), **self.get_dashboard_kwargs())
        ctx["period"] = self.request.GET.get("period") or "week"
        ctx["date_from"] = self.request.GET.get("from") or ""
        ctx["date_to"] = self.request.GET.get("to") or ""
        return ctx


class InvestDashboardExportView(InvestDashboardView):
    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import Font

        dashboard = build_dashboard(self.get_subsystem(), **self.get_dashboard_kwargs())
        wb = Workbook()
        ws = wb.active
        ws.title = "Funnel"
        ws.append(["Воронка", "Стадия", "Проектов"])
        for funnel_name, counts in (
            ("Привлечение", dashboard["attraction_counts"]),
            ("Сопровождение", dashboard["support_counts"]),
        ):
            for stage, total in counts.items():
                ws.append([funnel_name, stage, total])

        overdue_ws = wb.create_sheet("Overdue")
        overdue_ws.append(["Показатель", "Значение"])
        overdue_ws.append(["Просрочено шагов", dashboard["overdue_count"]])
        overdue_ws.append(["Готовность пакетов, %", dashboard["packages_ready_pct"]])
        overdue_ws.append(["Активные брони", dashboard["active_bookings"]])

        bottlenecks_ws = wb.create_sheet("Bottlenecks")
        bottlenecks_ws.append(["МО", "Просрочено"])
        for row in dashboard["bottlenecks_by_org"]:
            bottlenecks_ws.append([row["organization_name"], row["overdue_count"]])

        industry_ws = wb.create_sheet("Industry")
        industry_ws.append(["Отрасль", "Проектов", "Инвестиции, млн руб.", "Рабочие места"])
        for row in dashboard["industry_metrics"]:
            industry_ws.append([row["industry"], row["projects"], row["investment_amount"], row["jobs"]])

        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            for column in sheet.columns:
                letter = column[0].column_letter
                sheet.column_dimensions[letter].width = max(len(str(cell.value or "")) for cell in column) + 2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="invest-dashboard.xlsx"'
        wb.save(response)
        return response


class InvestCockpitView(InvestDashboardView):
    template_name = "invest/cockpit.html"
    page_title = "Executive cockpit"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        cockpit = build_cockpit(self.get_subsystem(), **self.get_dashboard_kwargs())
        ctx["cockpit"] = cockpit
        ctx["dashboard"] = cockpit["dashboard"]
        return ctx


class InvestDashboardBriefView(InvestDashboardView):
    def get(self, request, *args, **kwargs):
        cockpit = build_cockpit(self.get_subsystem(), **self.get_dashboard_kwargs())
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4, pageCompression=0)
        width, height = A4
        regular_font, bold_font = _passport_pdf_fonts()
        y = height - 48
        pdf.setFont(bold_font, 16)
        pdf.drawString(48, y, "Invest meeting brief")
        y -= 28
        pdf.setFont(regular_font, 10)
        rows = [
            ("Projects total", cockpit["kpis"]["projects_total"]),
            ("Attraction", cockpit["kpis"]["attraction"]),
            ("Support", cockpit["kpis"]["support"]),
            ("Overdue", cockpit["kpis"]["overdue"]),
            ("Packages ready pct", f"{cockpit['kpis']['packages_ready_pct']}%"),
            ("Active bookings", cockpit["kpis"]["active_bookings"]),
            ("SLA high", cockpit["sla_risk"]["high_count"]),
            ("SLA medium", cockpit["sla_risk"]["medium_count"]),
            (
                "Quarter target",
                f"{cockpit['quarter_target']['actual']} / {cockpit['quarter_target']['goal']}",
            ),
        ]
        for label, value in rows:
            pdf.setFont(bold_font, 10)
            pdf.drawString(48, y, f"{label}:")
            pdf.setFont(regular_font, 10)
            pdf.drawString(170, y, str(value))
            y -= 18

        y -= 10
        pdf.setFont(bold_font, 12)
        pdf.drawString(48, y, "Heat by MO")
        y -= 18
        pdf.setFont(regular_font, 10)
        for row in cockpit["heat_by_mo"][:8]:
            pdf.drawString(48, y, f"{row['organization_name']}: {row['overdue_count']}")
            y -= 16
        pdf.drawString(48, y, cockpit["heat_note"])
        pdf.showPage()
        pdf.save()
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="invest-meeting-brief.pdf"'
        return response


class InvestKanbanView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/kanban.html"
    page_title = "Канбан инвестпроектов"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        projects = projects_for_membership(self.get_membership()).select_related("organization", "owner")
        kanban = {}
        for funnel, _label in InvestProject.Funnel.choices:
            kanban[funnel] = {}
            for project in projects.filter(funnel=funnel).order_by("stage", "code"):
                column = kanban[funnel].setdefault(
                    project.stage,
                    {
                        "stage": project.stage,
                        "label": InvestProjectForm.STAGE_LABELS.get(project.stage, project.stage),
                        "projects": [],
                    },
                )
                column["projects"].append(project)
        ctx["kanban"] = kanban
        return ctx


class InvestInboxView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/inbox.html"
    page_title = "Сегодня"
    http_method_names = ["get", "post", "head", "options"]

    def post(self, request, *args, **kwargs):
        if request.POST.get("action") == "refresh_sla":
            result = refresh_invest_sla(
                subsystem=self.get_subsystem(),
                user=request.user,
                request=request,
            )
            messages.success(
                request,
                f"SLA обновлены: дорожная карта {result['roadmap']}, задачи {result['external_tasks']}.",
            )
        return redirect(reverse("invest-inbox"))

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        projects = projects_for_membership(membership)
        now = timezone.now()
        ctx["roadmap_items"] = (
            InvestRoadmapItem.objects.filter(project__in=projects)
            .filter(Q(status=InvestRoadmapItem.Status.OVERDUE) | Q(status=InvestRoadmapItem.Status.OPEN, due_at__lt=now))
            .select_related("project", "project__organization", "owner")
            .order_by("due_at", "code")[:50]
        )
        ctx["handoffs"] = (
            InvestHandoff.objects.filter(project__in=projects, status=InvestHandoff.Status.REQUESTED)
            .select_related("project", "project__organization", "requested_by")
            .order_by("created_at")[:50]
        )
        ctx["external_tasks"] = (
            InvestExternalTask.objects.filter(
                project__in=projects,
                status__in=(InvestExternalTask.Status.OPEN, InvestExternalTask.Status.OVERDUE),
            )
            .select_related("project", "organization")
            .order_by("due_at", "-created_at")[:50]
        )
        ctx["overdue_extracts"] = extracts_for_inbox(projects=projects, now=now)
        ctx["overdue_fgistp"] = fgistp_for_inbox(projects=projects, now=now)
        return ctx


class InvestProjectListView(InvestSubsystemMixin, ModulePermissionMixin, ListView):
    model = InvestProject
    template_name = "invest/projects_list.html"
    context_object_name = "projects"
    page_title = "Инвестпроекты"
    paginate_by = 25

    def get_queryset(self):
        return projects_for_membership(self.get_membership()).select_related("organization", "owner")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        ctx["can_create_project"] = user_can(self.request.user, self.module_code, "create")
        ctx["can_bulk_stage"] = user_can(self.request.user, self.module_code, "change") and _can_bulk_update_stage(
            self.request.user, membership
        )
        ctx["bulk_stage_choices"] = [
            (stage, InvestProjectForm.STAGE_LABELS.get(stage, stage))
            for stage in _stage_to_funnel()
        ]
        return ctx


class InvestProjectBulkStageView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        membership = self.get_membership()
        if not _can_bulk_update_stage(request.user, membership):
            return _forbidden_with_message(request, "Массовая смена стадии доступна только администратору или департаменту")

        project_ids = request.POST.getlist("project_ids")
        new_stage = (request.POST.get("stage") or request.POST.get("new_stage") or "").strip()
        target_funnel = _stage_to_funnel().get(new_stage)
        if not project_ids or not target_funnel:
            messages.error(request, "Выберите проекты и корректную стадию.")
            return redirect(reverse("invest-projects"))

        projects = list(projects_for_membership(membership).filter(pk__in=project_ids).select_related("organization"))
        if len(projects) != len(set(project_ids)):
            messages.error(request, "Часть проектов недоступна в текущем контуре.")
            return redirect(reverse("invest-projects"))
        if any(project.funnel != target_funnel for project in projects):
            messages.error(request, "Массовая смена стадии разрешена только внутри той же воронки.")
            return redirect(reverse("invest-projects"))

        with transaction.atomic():
            for project in projects:
                old_stage = project.stage
                if old_stage == new_stage:
                    continue
                project.stage = new_stage
                project.save(update_fields=["stage", "updated_at"])
                audit.log_action(
                    request.user,
                    membership.subsystem,
                    "invest.project.bulk_stage",
                    model_name="InvestProject",
                    object_id=project.pk,
                    payload={"old_stage": old_stage, "new_stage": new_stage},
                    request=request,
                )
        messages.success(request, f"Стадия обновлена для {len(projects)} проектов.")
        return redirect(reverse("invest-projects"))


class InvestProjectDetailView(InvestSubsystemMixin, ModulePermissionMixin, DetailView):
    model = InvestProject
    template_name = "invest/project_detail.html"
    context_object_name = "project"
    page_title = "Инвестпроект"

    def get_queryset(self):
        return projects_for_membership(self.get_membership()).select_related("organization", "owner")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        project = self.object
        package = ensure_package(project)
        items = list(package.items.order_by("id"))
        required = [i for i in items if i.required]
        ready = sum(1 for i in required if i.status == InvestPackageItem.Status.ATTACHED)
        ctx["can_change_project"] = user_can(self.request.user, self.module_code, "change")
        ctx["site_links"] = project.site_links.select_related("site", "site__organization").order_by("role", "id")
        ctx["roadmap_items"] = project.roadmap_items.select_related("owner").order_by("due_at", "code")
        ctx["support_track_items"] = project.support_track_items.order_by("due_at", "created_at")
        ctx["support_track_form"] = InvestSupportTrackItemForm()
        ctx["protocols"] = project.protocols.select_related("document").order_by("-signed_at", "-created_at")
        ctx["oiv_approvals"] = project.oiv_approvals.order_by("due_at", "agency_name")
        ctx["stop_factors"] = project.stop_factors.order_by("status", "created_at")
        ctx["open_stop_factors"] = _open_stop_factors(project)
        ctx["package"] = package
        ctx["package_items"] = items
        ctx["package_ready"] = f"{ready}/{len(required)}" if required else "—"
        blockers = gate_blockers(project)
        can_push, bitrix_blockers = can_push_to_bitrix(project)
        ctx["bitrix_can_push"] = can_push
        ctx["bitrix_blockers"] = bitrix_blockers
        ctx["bitrix_stage_conflict"] = (project.external_ids or {}).get("bitrix_stage_conflict")
        ctx["inn_check"] = {
            "inn": (project.external_ids or {}).get("investor_inn", ""),
            "valid": (project.external_ids or {}).get("investor_inn_valid"),
            "checked_at": (project.external_ids or {}).get("investor_inn_checked_at"),
        }
        ctx["completeness_pct"] = compute_completeness(project)
        ctx["completeness_actions"] = _completeness_actions(project, blockers)
        ctx["sla_risk"] = project_sla_risk(project)
        ctx["handoff_stop_factor_blocked"] = _open_stop_factors(project).exists()
        ctx["audit_logs"] = AuditLog.objects.filter(
            subsystem=project.subsystem,
            model_name__iexact="InvestProject",
            object_id=str(project.pk),
        )[:20]
        ctx["project_comments"] = project.comments.select_related("author")
        ctx["can_comment_project"] = user_can(self.request.user, self.module_code, "change")
        _with_opendata_context(
            ctx,
            entity=project,
            verify_url=reverse("invest-verification-project", args=[project.pk]),
            can_verify=ctx["can_change_project"],
        )
        return _with_odysseus_cta(self.request, ctx, membership=membership, project=project)


class InvestInvestorProjectView(InvestSubsystemMixin, ModulePermissionMixin, DetailView):
    model = InvestProject
    template_name = "invest/investor_project_view.html"
    context_object_name = "project"
    page_title = "Investor view"

    def get_queryset(self):
        return projects_for_membership(self.get_membership()).select_related("organization")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["limited_fields"] = [
            ("Код", self.object.code),
            ("Проект", self.object.name),
            ("Инвестор", self.object.investor_name or "—"),
            ("Отрасль", self.object.industry or "—"),
            ("Стадия", self.object.stage),
            ("Воронка", self.object.get_funnel_display()),
            ("МО / территория", self.object.organization.name),
            ("Инвестиции, млн руб.", self.object.investment_amount or "—"),
            ("Рабочие места", self.object.jobs_count or "—"),
        ]
        return ctx


class InvestProjectCommentAddView(InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        project = get_object_or_404(projects_for_membership(self.get_membership()), pk=kwargs["pk"])
        body = (request.POST.get("body") or "").strip()
        if body:
            InvestProjectComment.objects.create(project=project, author=request.user, body=body)
            messages.success(request, "Комментарий добавлен.")
        else:
            messages.warning(request, "Введите текст комментария.")
        return redirect(reverse("invest-project-detail", args=[project.pk]))


class InvestProjectInnCheckView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        project = get_object_or_404(projects_for_membership(self.get_membership()), pk=kwargs["pk"])
        raw_inn = ""
        if project.investor_entity_id:
            raw_inn = project.investor_entity.inn
        raw_inn = raw_inn or (project.external_ids or {}).get("investor_inn", "")
        normalized = re.sub(r"\D", "", raw_inn or "")
        valid = inn_is_valid(normalized)
        external_ids = dict(project.external_ids or {})
        external_ids.update(
            {
                "investor_inn": normalized,
                "investor_inn_valid": valid,
                "investor_inn_checked_at": timezone.now().isoformat(),
            }
        )
        project.external_ids = external_ids
        project.save(update_fields=["external_ids", "updated_at"])
        if valid:
            try:
                run = run_project_verification(project, user=request.user)
                hard = int((run.summary or {}).get("hard_count") or 0)
                messages.success(
                    request,
                    f"ИНН валиден. Проверка открытых данных: источников "
                    f"{(run.summary or {}).get('sources_ok', 0)}/{(run.summary or {}).get('sources_total', 0)}"
                    f", жёстких {hard}.",
                )
            except InvestOpenDataError as exc:
                messages.success(request, "ИНН проверен: контрольная сумма корректна.")
                messages.error(request, f"Открытые данные: {exc}")
        else:
            messages.error(request, "ИНН проверен: значение не прошло контрольную проверку.")
        return redirect(reverse("invest-project-detail", args=[project.pk]) + "#opendata-results")


class InvestProjectPassportView(InvestSubsystemMixin, ModulePermissionMixin, View):
    def get(self, request, *args, **kwargs):
        project = get_object_or_404(
            projects_for_membership(self.get_membership()).select_related("organization", "owner"),
            pk=kwargs["pk"],
        )
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        regular_font, bold_font = _passport_pdf_fonts()
        y = height - 48
        pdf.setFont(bold_font, 16)
        pdf.drawString(48, y, "Invest project passport")
        y -= 32
        pdf.setFont(regular_font, 10)

        rows = [
            ("Code", project.code),
            ("Name", project.name),
            ("Investor", project.investor_name or "-"),
            ("Stage", project.stage),
            ("Organization", project.organization.name if project.organization_id else "-"),
            (
                "Investments",
                str(project.investment_amount) if project.investment_amount is not None else "-",
            ),
            ("Jobs", str(project.jobs_count) if project.jobs_count is not None else "-"),
        ]
        for label, value in rows:
            pdf.setFont(bold_font, 10)
            pdf.drawString(48, y, f"{label}:")
            pdf.setFont(regular_font, 10)
            pdf.drawString(150, y, str(value))
            y -= 18

        y -= 8
        pdf.setFont(bold_font, 12)
        pdf.drawString(48, y, "Sites")
        y -= 18
        pdf.setFont(regular_font, 10)
        site_links = project.site_links.select_related("site").order_by("role", "id")
        if not site_links.exists():
            pdf.drawString(48, y, "-")
        for link in site_links:
            site = link.site
            summary = f"{site.cadastral_number} | {site.name} | {site.address or '-'} | {link.role}"
            pdf.drawString(48, y, summary[:110])
            y -= 16
            if y < 64:
                pdf.showPage()
                pdf.setFont(regular_font, 10)
                y = height - 48

        pdf.showPage()
        pdf.save()
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{project.code}-passport.pdf"'
        return response


class InvestProjectSupportAddView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        project = get_object_or_404(projects_for_membership(self.get_membership()), pk=kwargs["pk"])
        form = InvestSupportTrackItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.project = project
            item.save()
            messages.success(request, "Мера поддержки добавлена.")
        else:
            messages.error(request, "Проверьте поля меры поддержки.")
        return redirect(reverse("invest-project-detail", args=[project.pk]))


class InvestProjectBitrixPushView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        project = get_object_or_404(projects_for_membership(self.get_membership()), pk=kwargs["pk"])
        try:
            result = push_project_to_bitrix(project=project, force=False)
        except InvestBitrixError as exc:
            messages.error(request, str(exc))
        else:
            if result.get("pushed"):
                messages.success(request, "Проект отправлен в Bitrix.")
            else:
                messages.warning(request, "Отправка в Bitrix заблокирована: " + ", ".join(result.get("blockers") or []))
        return redirect(reverse("invest-project-detail", args=[project.pk]))


class InvestProjectBitrixConflictView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        project = get_object_or_404(projects_for_membership(self.get_membership()), pk=kwargs["pk"])
        resolution = request.POST.get("resolution") or ""
        try:
            result = resolve_bitrix_stage_conflict(project=project, resolution=resolution)
        except InvestBitrixError as exc:
            messages.error(request, str(exc))
        else:
            if result.get("resolved"):
                messages.success(request, "Конфликт стадий Bitrix/Delayu разрешён.")
            else:
                messages.info(request, "Активного конфликта стадий нет.")
        return redirect(reverse("invest-project-detail", args=[project.pk]))


class InvestInvestorListView(InvestSubsystemMixin, ModulePermissionMixin, ListView):
    model = InvestInvestor
    template_name = "invest/investors_list.html"
    context_object_name = "investors"
    page_title = "Юрлица инвесторов"
    paginate_by = 25

    def get_queryset(self):
        return (
            InvestInvestor.objects.filter(subsystem=self.get_subsystem())
            .prefetch_related("projects")
            .order_by("name")
        )


class InvestInvestorDetailView(InvestSubsystemMixin, ModulePermissionMixin, DetailView):
    model = InvestInvestor
    template_name = "invest/investor_detail.html"
    context_object_name = "investor"
    page_title = "Юрлицо инвестора"

    def get_queryset(self):
        return InvestInvestor.objects.filter(subsystem=self.get_subsystem())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        ctx["projects"] = projects_for_membership(membership).filter(investor_entity=self.object)
        _with_opendata_context(
            ctx,
            entity=self.object,
            verify_url=reverse("invest-verification-investor", args=[self.object.pk]),
            can_verify=user_can(self.request.user, self.module_code, "change"),
        )
        return ctx


class InvestVerificationInvestorView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        investor = get_object_or_404(
            InvestInvestor.objects.filter(subsystem=self.get_subsystem()),
            pk=kwargs["pk"],
        )
        try:
            run = run_investor_verification(investor, user=request.user)
            messages.success(
                request,
                f"Проверка открытых данных завершена: жёстких {(run.summary or {}).get('hard_count', 0)}.",
            )
        except InvestOpenDataError as exc:
            messages.error(request, str(exc))
        return redirect(reverse("invest-investor-detail", args=[investor.pk]) + "#opendata-results")


class InvestVerificationProjectView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        project = get_object_or_404(projects_for_membership(self.get_membership()), pk=kwargs["pk"])
        try:
            run = run_project_verification(project, user=request.user)
            messages.success(
                request,
                f"Проверка открытых данных завершена: жёстких {(run.summary or {}).get('hard_count', 0)}.",
            )
        except InvestOpenDataError as exc:
            messages.error(request, str(exc))
        return redirect(reverse("invest-project-detail", args=[project.pk]) + "#opendata-results")


class InvestVerificationSiteView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        site = get_object_or_404(sites_for_membership(self.get_membership()), pk=kwargs["pk"])
        try:
            run = run_site_verification(site, user=request.user)
            messages.success(
                request,
                f"Проверка площадки по открытым данным: жёстких {(run.summary or {}).get('hard_count', 0)}.",
            )
        except InvestOpenDataError as exc:
            messages.error(request, str(exc))
        return redirect(reverse("invest-site-detail", args=[site.pk]) + "#opendata-results")


class InvestOpenDataCatalogView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/opendata_catalog.html"
    page_title = "Открытые данные"

    def get_context_data(self, **kwargs):
        from delayu.services.invest_opendata import catalog_rows
        from delayu.models_invest import InvestVerificationRun

        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        flags = ensure_automation_config(membership.subsystem).get_flags()
        ctx["sources"] = catalog_rows()
        ctx["opendata_live"] = bool(flags.get("opendata_live") and not flags.get("opendata_mock", True))
        ctx["can_manage_flags"] = user_can(self.request.user, self.module_code, "change") and (
            getattr(membership.role, "code", "") == "invest_admin"
            or is_platform_admin(self.request.user)
        )
        ctx["recent_runs"] = (
            InvestVerificationRun.objects.filter(subsystem=membership.subsystem)
            .select_related("project", "investor", "site")
            .order_by("-created_at")[:20]
        )
        return ctx


class InvestDedupeView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/dedupe.html"
    page_title = "Дедупликация проектов"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["pairs"] = suspected_duplicate_pairs(self.get_subsystem())
        ctx["can_change_dedupe"] = user_can(self.request.user, self.module_code, "change")
        return ctx


class InvestDedupeIgnoreView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        membership = self.get_membership()
        projects = projects_for_membership(membership)
        left = get_object_or_404(projects, pk=request.POST.get("left_id"))
        right = get_object_or_404(projects, pk=request.POST.get("right_id"))
        ignore_duplicate_pair(left, right)
        messages.success(request, "Пара дублей скрыта.")
        return redirect(reverse("invest-dedupe"))


class InvestOdysseusOpenView(InvestSubsystemMixin, ModulePermissionMixin, View):
    """Prepare an Invest context snapshot before redirecting to Odysseus."""

    def get(self, request, *args, **kwargs):
        membership = self.get_membership()
        project = None
        site = None
        if request.GET.get("project"):
            project = get_object_or_404(projects_for_membership(membership), pk=request.GET["project"])
        if request.GET.get("site"):
            site = get_object_or_404(sites_for_membership(membership), pk=request.GET["site"])
        try:
            return redirect(
                prepare_odysseus_open(
                    request,
                    membership=membership,
                    project=project,
                    site=site,
                )
            )
        except PermissionError as exc:
            return HttpResponseForbidden(str(exc))

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)


class InvestHandoffListView(InvestSubsystemMixin, ModulePermissionMixin, ListView):
    model = InvestHandoff
    template_name = "invest/handoffs_list.html"
    context_object_name = "handoffs"
    page_title = "Передачи в сопровождение"
    paginate_by = 25

    def get_queryset(self):
        projects = projects_for_membership(self.get_membership())
        return (
            InvestHandoff.objects.filter(project__in=projects)
            .select_related("project", "project__organization", "requested_by", "decided_by")
            .order_by("-created_at")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["can_change_handoff"] = user_can(self.request.user, self.module_code, "change")
        ctx["return_reason_templates"] = RETURN_REASON_TEMPLATES
        return ctx


class InvestHandoffRequestView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"
    allowed_role_codes = {"invest_agency", "invest_admin"}

    def post(self, request, *args, **kwargs):
        membership = self.get_membership()
        if not _membership_has_role(membership, self.allowed_role_codes):
            return _forbidden_with_message(request, "Передачу может запросить только агентство")
        project = get_object_or_404(projects_for_membership(membership), pk=kwargs["pk"])
        try:
            request_handoff(project=project, user=request.user, comment=request.POST.get("comment", ""))
        except InvestHandoffError as exc:
            messages.error(request, str(exc))
        else:
            messages.success(request, "Передача запрошена.")
        return redirect(reverse("invest-handoffs"))


class InvestHandoffDecisionView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"
    allowed_role_codes = {"invest_dept", "invest_admin"}
    decision = ""
    success_message = ""

    def get_handoff(self):
        projects = projects_for_membership(self.get_membership())
        return get_object_or_404(InvestHandoff.objects.filter(project__in=projects), pk=self.kwargs["pk"])

    def post(self, request, *args, **kwargs):
        membership = self.get_membership()
        if not _membership_has_role(membership, self.allowed_role_codes):
            return _forbidden_with_message(request, "Решение по передаче доступно только департаменту")
        handoff = self.get_handoff()
        try:
            if self.decision == "accept":
                accept_handoff(handoff=handoff, user=request.user)
            else:
                return_handoff(
                    handoff=handoff,
                    user=request.user,
                    comment=resolve_return_comment(
                        template_key=request.POST.get("comment_template", ""),
                        comment=request.POST.get("comment", ""),
                        fallback=handoff.comment,
                    ),
                )
        except InvestHandoffError as exc:
            messages.error(request, str(exc))
        else:
            messages.success(request, self.success_message)
        return redirect(reverse("invest-handoffs"))


class InvestHandoffAcceptView(InvestHandoffDecisionView):
    decision = "accept"
    success_message = "Передача принята."


class InvestHandoffReturnView(InvestHandoffDecisionView):
    decision = "return"
    success_message = "Передача возвращена."


class InvestPackageDetailView(InvestSubsystemMixin, ModulePermissionMixin, DetailView):
    model = InvestProject
    template_name = "invest/package_detail.html"
    context_object_name = "project"
    page_title = "Пакет передачи"

    def get_queryset(self):
        return projects_for_membership(self.get_membership()).select_related("organization", "owner")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        package = ensure_package(self.object)
        ctx["package"] = package
        ctx["items"] = package.items.select_related("document").order_by("id")
        ctx["snapshots"] = package.snapshots.select_related("handoff").order_by("-created_at")
        ctx["status_choices"] = InvestPackageItem.Status.choices
        ctx["can_change_package"] = user_can(self.request.user, self.module_code, "change")
        ctx["recent_documents"] = DocumentFile.objects.filter(
            subsystem=self.get_subsystem(),
            is_current=True,
        ).order_by("-created_at")[:25]
        return ctx


class InvestPackageItemUpdateView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        project = get_object_or_404(projects_for_membership(self.get_membership()), pk=kwargs["project_pk"])
        package = ensure_package(project)
        item = get_object_or_404(package.items, pk=kwargs["item_pk"])
        status = request.POST.get("status")
        if status not in InvestPackageItem.Status.values:
            messages.error(request, "Некорректный статус пункта пакета.")
        else:
            document = None
            document_id = request.POST.get("document")
            if document_id:
                document = get_object_or_404(DocumentFile, pk=document_id, subsystem=self.get_subsystem())
            set_item_status(item, status, request.FILES.get("file"), document)
            messages.success(request, "Пункт пакета обновлён.")
        return redirect(reverse("invest-package-detail", args=[project.pk]))


class InvestImportListView(InvestImportRoleMixin, InvestSubsystemMixin, ModulePermissionMixin, ListView):
    model = InvestImportBatch
    template_name = "invest/import_list.html"
    context_object_name = "batches"
    page_title = "Импорт данных МО"
    paginate_by = 25
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        return _import_batches_for_membership(self.get_membership()).order_by("-created_at")

    def post(self, request, *args, **kwargs):
        membership = self.get_membership()
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Выберите CSV-файл для импорта.")
            return redirect(reverse("invest-imports"))

        batch = parse_mo_file(upload, subsystem=membership.subsystem, organization=membership.organization)
        messages.success(request, "Файл загружен. Проверьте строки перед применением.")
        return redirect(reverse("invest-import-detail", args=[batch.pk]))


class InvestImportDetailView(InvestImportRoleMixin, InvestSubsystemMixin, ModulePermissionMixin, DetailView):
    model = InvestImportBatch
    template_name = "invest/import_detail.html"
    context_object_name = "batch"
    page_title = "Пакет импорта"

    def get_queryset(self):
        return _import_batches_for_membership(self.get_membership())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["rows"] = self.object.rows.select_related("target_project", "target_site").order_by("row_number")
        return ctx


class InvestImportRowActionView(InvestImportRoleMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    action = ""

    def post(self, request, *args, **kwargs):
        batch = get_object_or_404(_import_batches_for_membership(self.get_membership()), pk=kwargs["batch_pk"])
        row = get_object_or_404(batch.rows, pk=kwargs["row_pk"])
        try:
            if self.action == "apply":
                apply_row(row, user=request.user)
                messages.success(request, "Строка применена.")
            else:
                skip_row(row)
                messages.success(request, "Строка пропущена.")
        except ValueError as exc:
            messages.error(request, str(exc))

        if not batch.rows.filter(resolution=InvestImportRow.Resolution.PENDING).exists():
            batch.status = InvestImportBatch.Status.DONE
            batch.save(update_fields=["status"])
        return redirect(reverse("invest-import-detail", args=[batch.pk]))


class InvestImportRowApplyView(InvestImportRowActionView):
    action = "apply"


class InvestImportRowSkipView(InvestImportRowActionView):
    action = "skip"


class InvestProjectWizardView(
    InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, ContextMixin, View
):
    """Multi-step View: ContextMixin needed so PlatformLayoutMixin can set layout_path."""

    template_name = "invest/project_wizard.html"
    required_action = "create"
    page_title = "Мастер нового инвестпроекта"

    def _step(self):
        raw = self.request.POST.get("step") or self.request.GET.get("step") or "1"
        return raw if raw in {"1", "2", "3"} else "1"

    def _session_data(self):
        return dict(self.request.session.get(PROJECT_WIZARD_SESSION_KEY) or {})

    def _save_session_data(self, data):
        self.request.session[PROJECT_WIZARD_SESSION_KEY] = data
        self.request.session.modified = True

    def _form_for_step(self, step, data=None):
        membership = self.get_membership()
        initial = self._session_data()
        if step == "1":
            return InvestProjectWizardStep1Form(data=data, initial=initial.get("step1"), membership=membership)
        if step == "2":
            return InvestProjectWizardStep2Form(data=data, initial=initial.get("step2"), membership=membership)
        return InvestProjectWizardStep3Form(data=data, initial=initial.get("step3"))

    def _render(self, step, form):
        return render(
            self.request,
            self.template_name,
            self.get_context_data(
                form=form,
                step=step,
                step_number=int(step),
                wizard_data=self._session_data(),
            ),
        )

    def get(self, request, *args, **kwargs):
        step = self._step()
        if step != "1" and not self._session_data().get("step1"):
            return redirect(reverse("invest-project-wizard"))
        return self._render(step, self._form_for_step(step))

    def post(self, request, *args, **kwargs):
        step = self._step()
        if step != "1" and not self._session_data().get("step1"):
            return redirect(reverse("invest-project-wizard"))

        form = self._form_for_step(step, data=request.POST)
        if not form.is_valid():
            return self._render(step, form)

        data = self._session_data()
        if step == "1":
            organization = form.cleaned_data["organization"]
            data["step1"] = {
                "organization": organization.pk,
                "code": form.cleaned_data["code"],
                "name": form.cleaned_data["name"],
                "investor_name": form.cleaned_data["investor_name"],
            }
            self._save_session_data(data)
            return redirect(f"{reverse('invest-project-wizard')}?step=2")

        if step == "2":
            site = form.cleaned_data.get("site")
            data["step2"] = {"site": site.pk if site else ""}
            self._save_session_data(data)
            return redirect(f"{reverse('invest-project-wizard')}?step=3")

        project = self._create_project_from_session(form.cleaned_data)
        request.session.pop(PROJECT_WIZARD_SESSION_KEY, None)
        request.session.modified = True
        if flag_enabled(project.subsystem, "auto_smev"):
            for link in project.site_links.select_related("site"):
                auto_smev_enrich_site(link.site, user=request.user)
        messages.success(request, "Инвестпроект создан через мастер.")
        return redirect(reverse("invest-project-detail", args=[project.pk]))

    def _create_project_from_session(self, cleaned_step3):
        membership = self.get_membership()
        data = self._session_data()
        step1 = data["step1"]
        project = InvestProject.objects.create(
            subsystem=membership.subsystem,
            organization_id=step1["organization"],
            code=step1["code"],
            name=step1["name"],
            investor_name=step1.get("investor_name", ""),
            funnel=InvestProject.Funnel.ATTRACTION,
            stage="lead",
            owner=self.request.user,
            municipality_notes=cleaned_step3.get("package_note", ""),
        )
        site_id = (data.get("step2") or {}).get("site")
        if site_id:
            site = get_object_or_404(sites_for_membership(membership), pk=site_id)
            InvestProjectSite.objects.create(
                project=project,
                site=site,
                role=InvestProjectSite.Role.PROPOSED,
            )
        return project


class InvestProjectCreateView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, CreateView):
    model = InvestProject
    form_class = InvestProjectForm
    template_name = "invest/project_form.html"
    page_title = "Новый инвестпроект"
    required_action = "create"
    success_url = reverse_lazy("invest-projects")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["okved_industries"] = OKVED_INDUSTRIES
        return ctx

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["membership"] = self.get_membership()
        return kwargs

    def form_valid(self, form):
        form.instance.subsystem = self.get_subsystem()
        form.instance.funnel = InvestProject.Funnel.ATTRACTION
        membership = self.get_membership()
        if membership.role.code == "invest_mo":
            form.instance.organization = membership.organization
        if form.instance.owner_id is None:
            form.instance.owner = self.request.user
        messages.success(self.request, "Инвестпроект создан.")
        return super().form_valid(form)


class InvestProjectUpdateView(InvestSubsystemMixin, ModulePermissionMixin, UpdateView):
    model = InvestProject
    form_class = InvestProjectForm
    template_name = "invest/project_form.html"
    context_object_name = "project"
    page_title = "Редактирование инвестпроекта"
    required_action = "change"

    def get_queryset(self):
        return projects_for_membership(self.get_membership()).select_related("organization", "owner")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["membership"] = self.get_membership()
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["okved_industries"] = OKVED_INDUSTRIES
        return ctx

    def get_success_url(self):
        return reverse_lazy("invest-project-detail", args=[self.object.pk])

    def form_valid(self, form):
        messages.success(self.request, "Инвестпроект обновлён.")
        return super().form_valid(form)


class InvestSiteListView(InvestSubsystemMixin, ModulePermissionMixin, ListView):
    model = InvestSite
    template_name = "invest/sites_list.html"
    context_object_name = "sites"
    page_title = "Инвестплощадки"
    paginate_by = 25

    def get_queryset(self):
        return sites_for_membership(self.get_membership()).select_related("organization")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["can_create_site"] = user_can(self.request.user, self.module_code, "create")
        ctx["can_change_site"] = user_can(self.request.user, self.module_code, "change")
        return ctx


class InvestSiteCampaignView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/sites_campaign.html"
    page_title = "Кампании актуализации МО"
    required_action = "change"
    http_method_names = ["get", "post", "head", "options"]

    def _organizations(self):
        membership = self.get_membership()
        qs = membership.subsystem.organizations.filter(is_active=True).order_by("name")
        if membership.role.code == "invest_mo":
            qs = qs.filter(pk=membership.organization_id)
        return qs

    def post(self, request, *args, **kwargs):
        membership = self.get_membership()
        organization = get_object_or_404(self._organizations(), pk=request.POST.get("organization"))
        campaign_id = f"mo-refresh-{organization.pk}-{timezone.now().strftime('%Y%m%d%H%M%S')}"
        requested_at = timezone.now().isoformat()
        sites = list(sites_for_membership(membership).filter(organization=organization))
        for site in sites:
            external_ids = dict(site.external_ids or {})
            external_ids["mo_refresh"] = {
                "campaign_id": campaign_id,
                "status": "queued",
                "requested_at": requested_at,
                "organization_id": organization.pk,
            }
            site.external_ids = external_ids
            site.save(update_fields=["external_ids", "updated_at"])
        messages.success(request, f"Кампания актуализации создана: {len(sites)} площадок.")
        return redirect(reverse("invest-sites-campaign"))

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        sites = list(sites_for_membership(self.get_membership()).select_related("organization"))
        ctx["organizations"] = self._organizations()
        ctx["campaigns"] = _campaign_progress(sites)
        return ctx


class InvestSiteMapView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/sites_map.html"
    page_title = "Карта инвестплощадок"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        sites = (
            sites_for_membership(self.get_membership())
            .select_related("organization")
            .prefetch_related("project_links__project")
        )
        org_filter = (self.request.GET.get("org") or "").strip()
        if org_filter:
            if org_filter.isdigit():
                sites = sites.filter(organization_id=int(org_filter))
            else:
                sites = sites.filter(organization__code=org_filter)
        booking_filter = (self.request.GET.get("booking") or "").strip()
        if booking_filter:
            sites = sites.filter(project_links__role=booking_filter).distinct()
        mapped_sites = []
        map_points = []
        for site in sites:
            lat, lon = _coordinates_for_site(site)
            booking_info = _booking_info_for_site(site)
            booking_badge = booking_info.get("badge") or ""
            zones = _restriction_zones_for_map(site)
            has_coordinates = bool(lat and lon)
            mapped_sites.append({
                "site": site,
                "latitude": lat,
                "longitude": lon,
                "has_coordinates": has_coordinates,
                "booking_badge": booking_badge,
                "restriction_zones": zones,
            })
            if has_coordinates:
                map_extract = latest_map_extract(site)
                map_fgistp = latest_map_fgistp(site)
                map_points.append({
                    "lat": float(lat),
                    "lon": float(lon),
                    "name": site.name,
                    "cadastral": site.cadastral_number,
                    "url": reverse("invest-site-detail", args=[site.pk]),
                    "orgName": site.organization.name if site.organization_id else "",
                    "statusLabel": site.get_status_display(),
                    "areaHa": str(site.area_ha) if site.area_ha is not None else "",
                    "bookingBadge": booking_badge,
                    "projectName": booking_info.get("projectName") or "",
                    "projectCode": booking_info.get("projectCode") or "",
                    "projectUrl": booking_info.get("projectUrl") or "",
                    "restrictionZones": zones,
                    "zoneCount": len(zones),
                    "hasExtract": bool(map_extract and getattr(map_extract, "geometry", None)),
                    "hasFgistp": bool(map_fgistp and getattr(map_fgistp, "geometry", None)),
                    "smevRgisIntersections": (site.external_ids or {}).get("smev_rgis_intersections") or "",
                    "extractGeometry": extract_geometry_for_map(map_extract),
                    "fgistpGeometry": fgistp_geometry_for_map(map_fgistp),
                })
        ctx["mapped_sites"] = mapped_sites
        ctx["map_points_json"] = json.dumps(map_points, cls=DjangoJSONEncoder, ensure_ascii=False)
        ctx["has_coordinates"] = any(entry["has_coordinates"] for entry in mapped_sites)
        mnp_cfg = mnp_map_config(self.get_membership().subsystem)
        mnp_store = store_status()
        ctx["mnp_config"] = mnp_cfg
        ctx["mnp_store"] = mnp_store
        ctx["mnp_config_json"] = json.dumps(
            {
                "tileUrlTemplate": reverse(
                    "invest-mnp-tile", kwargs={"z": 99, "x": 88, "y": 77}
                ).replace("/99/88/77.png", "/{z}/{x}/{y}.png"),
                "viewportProxy": reverse("invest-mnp-viewport"),
                "tileZmax": int(getattr(settings, "INVEST_MNP_TILE_ZMAX", 12)),
                "vectorZoom": int(getattr(settings, "INVEST_MNP_VECTOR_ZOOM", 11)),
                "detailZoom": int(getattr(settings, "INVEST_MNP_DETAIL_ZOOM", 12)),
                "viewportMode": str(getattr(settings, "INVEST_MNP_VIEWPORT_MODE", "auto") or "auto"),
                "liveWms": bool(getattr(settings, "INVEST_MNP_LIVE_WMS", True)),
                "styles": map_style_config(),
                "featuresProxy": reverse("invest-mnp-features"),
                "wmsProxy": reverse("invest-mnp-wms"),
                "wfsProxy": reverse("invest-mnp-wfs"),
                "viewerUrl": mnp_cfg["viewer_url"],
                "enabled": mnp_cfg["enabled"],
                "maxFeatures": mnp_cfg["max_features"],
                "store": mnp_store,
            },
            ensure_ascii=False,
        )
        return ctx


class InvestMnpTileView(InvestSubsystemMixin, ModulePermissionMixin, View):
    """Serve local MNP PNG tile (no upstream)."""

    def get(self, request, z: int, x: int, y: int, *args, **kwargs):
        try:
            z_i, x_i, y_i = int(z), int(x), int(y)
        except (TypeError, ValueError):
            return HttpResponse(status=400)
        if z_i < 0 or z_i > 18 or x_i < 0 or y_i < 0:
            return HttpResponse(status=400)
        content = read_tile_bytes(z_i, x_i, y_i)
        response = HttpResponse(content, content_type="image/png")
        # Avoid sticky browser cache of empty/transparent placeholders.
        response["Cache-Control"] = "public, max-age=120, must-revalidate"
        return response


class InvestMnpViewportView(InvestSubsystemMixin, ModulePermissionMixin, View):
    """Compose local tiles / GeoJSON / live WMS for map viewport."""

    def get(self, request, *args, **kwargs):
        try:
            content = render_viewport_png(
                bbox=request.GET.get("bbox") or "",
                width=int(request.GET.get("width") or 256),
                height=int(request.GET.get("height") or 256),
                mode=request.GET.get("mode"),
                zoom=request.GET.get("zoom"),
                allow_live=(
                    None
                    if request.GET.get("live") in (None, "")
                    else str(request.GET.get("live")).lower() in ("1", "true", "yes", "on")
                ),
            )
        except (InvestMnpError, ValueError, TypeError) as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        response = HttpResponse(content, content_type="image/png")
        response["Cache-Control"] = "public, max-age=300, must-revalidate"
        return response


class InvestMnpFeaturesView(InvestSubsystemMixin, ModulePermissionMixin, View):
    """GeoJSON features from local MNP store."""

    def get(self, request, *args, **kwargs):
        try:
            data = query_features_geojson(
                bbox=request.GET.get("bbox") or "",
                max_features=request.GET.get("max_features"),
            )
        except InvestMnpError as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        return JsonResponse(data, json_dumps_params={"ensure_ascii": False})


class InvestMnpWmsProxyView(InvestSubsystemMixin, ModulePermissionMixin, View):
    """Compat: bbox image from local tile mosaic / vector / live."""

    def get(self, request, *args, **kwargs):
        try:
            content = render_viewport_png(
                bbox=request.GET.get("bbox") or "",
                width=int(request.GET.get("width") or 256),
                height=int(request.GET.get("height") or 256),
                mode=request.GET.get("mode"),
                zoom=request.GET.get("zoom"),
            )
        except (InvestMnpError, ValueError, TypeError) as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        response = HttpResponse(content, content_type="image/png")
        response["Cache-Control"] = "public, max-age=60, must-revalidate"
        return response


class InvestMnpWfsProxyView(InvestSubsystemMixin, ModulePermissionMixin, View):
    """Compat: alias to local features store."""

    def get(self, request, *args, **kwargs):
        try:
            data = fetch_mnp_wfs(
                bbox=request.GET.get("bbox") or "",
                subsystem=self.get_subsystem(),
                max_features=request.GET.get("max_features"),
            )
        except InvestMnpError as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        return JsonResponse(data, json_dumps_params={"ensure_ascii": False})


class InvestSiteCompareView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/sites_compare.html"
    page_title = "Сравнение площадок"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        raw_ids = (self.request.GET.get("ids") or "").replace(";", ",").split(",")
        ids = []
        for raw_id in raw_ids:
            if raw_id.strip().isdigit():
                ids.append(int(raw_id.strip()))
            if len(ids) == 3:
                break
        sites = list(
            sites_for_membership(self.get_membership())
            .filter(pk__in=ids)
            .select_related("organization")
        )
        site_by_id = {site.pk: site for site in sites}
        ctx["sites"] = [site_by_id[site_id] for site_id in ids if site_id in site_by_id]
        ctx["requested_ids"] = ",".join(str(site_id) for site_id in ids)
        return ctx


class InvestBookingsView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, ListView):
    model = InvestProjectSite
    template_name = "invest/bookings.html"
    context_object_name = "bookings"
    page_title = "Брони площадок"
    required_action = "change"
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        membership = self.get_membership()
        return (
            InvestProjectSite.objects.filter(
                project__in=projects_for_membership(membership),
                site__in=sites_for_membership(membership),
                role__in=(InvestProjectSite.Role.BOOKED, InvestProjectSite.Role.SELECTED),
            )
            .select_related("project", "project__organization", "site", "site__organization")
            .order_by("booked_until", "project__code", "site__cadastral_number")
        )

    def post(self, request, *args, **kwargs):
        if request.POST.get("action") == "expire":
            expired = expire_overdue_bookings(subsystem=self.get_subsystem())
            messages.success(request, f"Просроченные брони сняты: {expired}.")
        return redirect(reverse("invest-bookings"))


class InvestSiteDetailView(InvestSubsystemMixin, ModulePermissionMixin, DetailView):
    model = InvestSite
    template_name = "invest/site_detail.html"
    context_object_name = "site"
    page_title = "Инвестплощадка"

    def get_queryset(self):
        return sites_for_membership(self.get_membership()).select_related("organization")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        ctx["can_change_site"] = user_can(self.request.user, self.module_code, "change")
        ctx["can_booking_override"] = _can_admin_override(self.request.user, membership)
        ctx["project_links"] = self.object.project_links.select_related("project", "project__organization")
        ctx["projects"] = projects_for_membership(membership).select_related("organization").order_by("code")
        ctx["smev_requests"] = self.object.smev_requests.all()[:10]
        flags = ensure_automation_config(self.object.subsystem).get_flags()
        ctx["smev_live_pending_mode"] = bool(flags.get("smev_live") and not flags.get("smev_mock"))
        egrn_requests = self.object.smev_requests.filter(service=InvestSmevRequest.Service.EGRN)[:10]
        ctx["egrn_history"] = [
            {
                "request": request,
                "payload_summary": _smev_payload_summary(request.response_payload),
                "field_diff": _smev_field_diff(request.response_payload),
            }
            for request in egrn_requests
        ]
        ctx["smev_services"] = InvestSmevRequest.Service.choices
        ctx["smev_apply_fields"] = [
            "address",
            "land_category",
            "vri",
            "right_type",
            "encumbrances",
            "zone_info",
            "area_ha",
            "latitude",
            "longitude",
        ]
        ctx["smev_rgis_intersections"] = (self.object.external_ids or {}).get("smev_rgis_intersections")
        ctx["extracts"] = self.object.extracts.select_related("project", "requested_by", "verified_by").all()[:20]
        ctx["extract_map"] = extract_geometry_for_map(latest_map_extract(self.object))
        ctx["fgistp_records"] = self.object.fgistp_records.select_related("project", "requested_by", "verified_by").all()[:20]
        ctx["fgistp_map"] = fgistp_geometry_for_map(latest_map_fgistp(self.object))
        _with_opendata_context(
            ctx,
            entity=self.object,
            verify_url=reverse("invest-verification-site", args=[self.object.pk]),
            can_verify=ctx["can_change_site"],
        )
        return _with_odysseus_cta(self.request, ctx, membership=membership, site=self.object)


class InvestSiteSmevBatchView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    """Создаёт mock-запросы СМЭВ для списка кадастровых номеров."""

    required_action = "change"

    def post(self, request, *args, **kwargs):
        membership = self.get_membership()
        raw_numbers = request.POST.get("cadastral_numbers") or request.POST.get("cadastral_numbers[]") or ""
        requested = [item.strip() for item in re.split(r"[\s,;]+", raw_numbers) if item.strip()]
        if not requested:
            messages.warning(request, "СМЭВ batch: укажите хотя бы один кадастровый номер.")
            return redirect(reverse("invest-sites"))

        sites = list(
            sites_for_membership(membership)
            .filter(cadastral_number__in=requested)
            .select_related("organization")
        )
        summary = batch_smev_requests(sites=sites, user=request.user)
        missing_count = max(0, len(set(requested)) - len({site.cadastral_number for site in sites}))
        messages.success(
            request,
            f"СМЭВ batch {summary['batch_id']}: создано {summary['total']}, "
            f"готово {summary['done']}, ожидает {summary['pending']}, ошибок {summary['error']}; "
            f"пропущено кадастров: {missing_count}.",
        )
        return redirect(f"{reverse('invest-smev-console')}?batch_id={summary['batch_id']}&correlation_id={summary['correlation_id']}")


class InvestSiteCreateView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, CreateView):
    model = InvestSite
    form_class = InvestSiteForm
    template_name = "invest/site_form.html"
    page_title = "Новая инвестплощадка"
    required_action = "create"
    success_url = reverse_lazy("invest-sites")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["membership"] = self.get_membership()
        return kwargs

    def form_valid(self, form):
        form.instance.subsystem = self.get_subsystem()
        membership = self.get_membership()
        if membership.role.code == "invest_mo":
            form.instance.organization = membership.organization
        messages.success(self.request, "Инвестплощадка создана.")
        response = super().form_valid(form)
        if flag_enabled(self.object.subsystem, "auto_extract", default=True):
            ensure_extract_for_site(self.object, reason="site_created", user=self.request.user)
        if flag_enabled(self.object.subsystem, "auto_fgistp", default=True):
            ensure_fgistp_for_site(self.object, reason="site_created", user=self.request.user)
        return response


class InvestSiteUpdateView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, UpdateView):
    model = InvestSite
    form_class = InvestSiteForm
    template_name = "invest/site_form.html"
    page_title = "Редактирование площадки"
    required_action = "change"

    def get_queryset(self):
        return sites_for_membership(self.get_membership())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["membership"] = self.get_membership()
        return kwargs

    def get_success_url(self):
        return reverse("invest-site-detail", args=[self.object.pk])

    def post(self, request, *args, **kwargs):
        if request.POST.get("action") == "geocode":
            self.object = self.get_object()
            form = self.get_form()
            if not form.is_valid():
                return self.form_invalid(form)
            site = form.save(commit=False)
            membership = self.get_membership()
            if membership.role.code == "invest_mo":
                site.organization = membership.organization
            try:
                lat, lon = geocode_address(site.address or site.name or site.cadastral_number)
            except YandexGeocodeError as exc:
                messages.error(request, str(exc))
                return self.form_invalid(form)
            site.latitude = lat
            site.longitude = lon
            site.save()
            form.save_m2m()
            self.object = site
            messages.success(request, "Координаты площадки обновлены через Яндекс Геокодер.")
            return redirect(self.get_success_url())
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        membership = self.get_membership()
        if membership.role.code == "invest_mo":
            form.instance.organization = membership.organization
        messages.success(self.request, "Площадка сохранена.")
        return super().form_valid(form)


class InvestSiteSmevRequestView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    """Тестовый запрос СМЭВ по площадке."""

    required_action = "change"

    def post(self, request, *args, **kwargs):
        site = get_object_or_404(sites_for_membership(self.get_membership()), pk=kwargs["pk"])
        service = request.POST.get("service") or InvestSmevRequest.Service.EGRN
        req = request_smev_fill(site=site, user=request.user, service=service)
        if req.status == InvestSmevRequest.Status.LIVE_PENDING:
            messages.info(request, f"СМЭВ live ({req.get_service_display()}): запрос ожидает ответа шлюза.")
        else:
            messages.success(
                request,
                f"Тестовый СМЭВ ({req.get_service_display()}): ответ получен. Можно применить к карточке.",
            )
        return redirect(reverse("invest-site-detail", args=[site.pk]))


class InvestSiteSmevContourView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        site = get_object_or_404(sites_for_membership(self.get_membership()), pk=kwargs["pk"])
        created = request_contour_check(site=site, user=request.user)
        messages.success(request, f"Контурная проверка СМЭВ: создано запросов {len(created)}.")
        return redirect(reverse("invest-site-detail", args=[site.pk]))


class InvestSiteSmevEmulateView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        site = get_object_or_404(sites_for_membership(self.get_membership()), pk=kwargs["pk"])
        smev_req = get_object_or_404(InvestSmevRequest, pk=kwargs["request_pk"], site=site)
        emulate_gateway_response(smev_req, actor=request.user)
        messages.success(request, "Эмулятор шлюза: ответ получен.")
        return redirect(reverse("invest-site-detail", args=[site.pk]))


class InvestSiteSmevApplyView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        membership = self.get_membership()
        site = get_object_or_404(sites_for_membership(membership), pk=kwargs["pk"])
        smev_req = get_object_or_404(InvestSmevRequest, pk=kwargs["request_pk"], site=site)
        if not smev_req.is_mock and not user_can_apply_live(user=request.user, membership=membership):
            return HttpResponseForbidden("Live-применение доступно только invest_admin / superuser")
        fields = request.POST.getlist("fields") or None
        rejected = request.POST.getlist("rejected_fields") or None
        try:
            apply_smev_response(
                request=smev_req,
                user=request.user,
                fields=fields,
                rejected_fields=rejected,
            )
        except InvestSmevError as exc:
            messages.error(request, str(exc))
        else:
            messages.success(request, "Данные тестового СМЭВ применены к карточке площадки.")
        return redirect(reverse("invest-site-detail", args=[site.pk]))


class InvestSmevConsoleView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/smev_console.html"
    page_title = "СМЭВ — операторский пульт"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        qs = InvestSmevRequest.objects.filter(subsystem=membership.subsystem).select_related(
            "site", "site__organization", "created_by"
        )
        status = self.request.GET.get("status") or ""
        service = self.request.GET.get("service") or ""
        org = self.request.GET.get("org") or ""
        batch_id = self.request.GET.get("batch_id") or ""
        correlation_id = self.request.GET.get("correlation_id") or ""
        if status:
            qs = qs.filter(status=status)
        if service:
            qs = qs.filter(service=service)
        if org:
            qs = qs.filter(site__organization_id=org)
        if correlation_id:
            qs = qs.filter(correlation_id=correlation_id)
        elif batch_id:
            qs = qs.filter(correlation_id__icontains=batch_id)
        flags = ensure_automation_config(membership.subsystem).get_flags()
        ctx["smev_mode_label"] = "Live" if flags.get("smev_live") and not flags.get("smev_mock") else "Sandbox"
        ctx["requests"] = qs.order_by("-created_at")[:200]
        ctx["status_choices"] = InvestSmevRequest.Status.choices
        ctx["service_choices"] = InvestSmevRequest.Service.choices
        ctx["orgs"] = membership.subsystem.organizations.order_by("name")
        ctx["filter_status"] = status
        ctx["filter_service"] = service
        ctx["filter_org"] = org
        ctx["batch_id"] = batch_id
        ctx["correlation_id"] = correlation_id
        return ctx


class InvestSmevReportView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/smev_report.html"
    page_title = "СМЭВ — отчёт"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["report"] = build_smev_report(subsystem=self.get_subsystem(), days=7)
        return ctx


class InvestSmevProtocolView(InvestSubsystemMixin, ModulePermissionMixin, View):
    def get(self, request, *args, **kwargs):
        membership = self.get_membership()
        smev_req = get_object_or_404(
            InvestSmevRequest.objects.filter(subsystem=membership.subsystem),
            pk=kwargs["pk"],
        )
        return render_smev_protocol_pdf(smev_req)


class InvestSiteActionView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"
    action_service = None
    success_message = ""
    allow_booking_override = False

    def post(self, request, *args, **kwargs):
        membership = self.get_membership()
        project = get_object_or_404(projects_for_membership(membership), pk=kwargs["project_pk"])
        site = get_object_or_404(sites_for_membership(membership), pk=kwargs["site_pk"])
        service_kwargs = {"project": project, "site": site, "user": request.user}
        if self.allow_booking_override:
            service_kwargs["override_completeness_gate"] = bool(
                request.POST.get("booking_override") and _can_admin_override(request.user, membership)
            )
        try:
            self.action_service(**service_kwargs)
            if flag_enabled(site.subsystem, "auto_smev"):
                auto_smev_enrich_site(site, user=request.user)
        except InvestBookingError as exc:
            messages.error(request, str(exc))
        else:
            messages.success(request, self.success_message)
        return redirect(reverse("invest-site-detail", args=[site.pk]))


class InvestSiteBookView(InvestSiteActionView):
    action_service = staticmethod(book_site)
    success_message = "Площадка забронирована."
    allow_booking_override = True


class InvestSiteSelectView(InvestSiteActionView):
    action_service = staticmethod(select_site)
    success_message = "Площадка выбрана для проекта."


class InvestExtractListView(InvestSubsystemMixin, ModulePermissionMixin, ListView):
    model = InvestExtract
    template_name = "invest/extracts_list.html"
    context_object_name = "extracts"
    page_title = "Выкопировки"
    paginate_by = 50

    def get_queryset(self):
        membership = self.get_membership()
        qs = (
            InvestExtract.objects.filter(subsystem=membership.subsystem, site__in=sites_for_membership(membership))
            .select_related("site", "site__organization", "project", "requested_by")
            .order_by("-updated_at")
        )
        status = (self.request.GET.get("status") or "").strip()
        if status:
            qs = qs.filter(status=status)
        org = (self.request.GET.get("org") or "").strip()
        if org.isdigit():
            qs = qs.filter(site__organization_id=int(org))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        ctx["status_choices"] = InvestExtract.Status.choices
        ctx["filter_status"] = self.request.GET.get("status") or ""
        ctx["filter_org"] = self.request.GET.get("org") or ""
        from delayu.models import Organization

        ctx["orgs"] = Organization.objects.filter(subsystem=membership.subsystem).order_by("name")
        ctx["can_change"] = user_can(self.request.user, self.module_code, "change")
        return ctx


class InvestExtractDetailView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, DetailView):
    model = InvestExtract
    template_name = "invest/extract_detail.html"
    context_object_name = "extract"
    page_title = "Выкопировка"
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        membership = self.get_membership()
        return InvestExtract.objects.filter(
            subsystem=membership.subsystem,
            site__in=sites_for_membership(membership),
        ).select_related("site", "site__organization", "project")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["can_change"] = user_can(self.request.user, self.module_code, "change")
        geom = extract_geometry_for_map(self.object if self.object.geometry else None)
        ctx["extract_map"] = geom
        ctx["extract_map_json"] = json.dumps(geom, cls=DjangoJSONEncoder, ensure_ascii=False) if geom else "null"
        ctx["type_choices"] = InvestExtract.ExtractType.choices
        mnp_snap = intersections_from_extract(self.object)
        ctx["mnp_intersections"] = mnp_snap
        ctx["mnp_intersections_json"] = json.dumps(mnp_snap, cls=DjangoJSONEncoder, ensure_ascii=False)
        mnp_cfg = mnp_map_config(self.get_membership().subsystem)
        mnp_store = store_status()
        ctx["mnp_config"] = mnp_cfg
        ctx["mnp_store"] = mnp_store
        ctx["mnp_config_json"] = json.dumps(
            {
                "tileUrlTemplate": reverse(
                    "invest-mnp-tile", kwargs={"z": 99, "x": 88, "y": 77}
                ).replace("/99/88/77.png", "/{z}/{x}/{y}.png"),
                "viewportProxy": reverse("invest-mnp-viewport"),
                "tileZmax": int(getattr(settings, "INVEST_MNP_TILE_ZMAX", 12)),
                "vectorZoom": int(getattr(settings, "INVEST_MNP_VECTOR_ZOOM", 11)),
                "detailZoom": int(getattr(settings, "INVEST_MNP_DETAIL_ZOOM", 12)),
                "viewportMode": str(getattr(settings, "INVEST_MNP_VIEWPORT_MODE", "auto") or "auto"),
                "liveWms": bool(getattr(settings, "INVEST_MNP_LIVE_WMS", True)),
                "styles": map_style_config(),
                "featuresProxy": reverse("invest-mnp-features"),
                "wmsProxy": reverse("invest-mnp-wms"),
                "wfsProxy": reverse("invest-mnp-wfs"),
                "viewerUrl": mnp_cfg["viewer_url"],
                "enabled": mnp_cfg["enabled"],
                "maxFeatures": mnp_cfg["max_features"],
                "store": mnp_store,
            },
            ensure_ascii=False,
        )
        return ctx

    def post(self, request, *args, **kwargs):
        if not user_can(request.user, self.module_code, "change"):
            return _forbidden_with_message(request, "Недостаточно прав")
        self.object = self.get_object()
        action = (request.POST.get("action") or "").strip()
        try:
            if action == "request":
                ensure_extract_for_site(
                    self.object.site,
                    reason="manual",
                    user=request.user,
                    project=self.object.project,
                    force=True,
                )
                messages.success(request, "Запрос выкопировки создан/обновлён.")
            elif action == "upload":
                uploaded = request.FILES.get("file")
                if not uploaded:
                    messages.error(request, "Выберите файл.")
                else:
                    mark_extract_received(self.object, user=request.user, uploaded_file=uploaded)
                    messages.success(request, "Файл загружен, статус: получена.")
            elif action == "mock":
                generate_mock_contour(self.object, user=request.user)
                messages.success(request, "Mock-контур сгенерирован.")
            elif action == "import_geometry":
                uploaded = request.FILES.get("geometry_file")
                raw = request.POST.get("geometry_json") or ""
                if uploaded:
                    import_extract_geometry(
                        self.object,
                        raw=uploaded.read(),
                        filename=uploaded.name,
                        user=request.user,
                    )
                elif raw.strip():
                    import_extract_geometry(self.object, raw=raw, filename="inline.geojson", user=request.user)
                else:
                    raise InvestExtractError("Укажите GeoJSON или файл")
                messages.success(request, "Геометрия импортирована.")
            elif action == "verify":
                verify_extract(self.object, user=request.user, attach=True)
                messages.success(request, "Выкопировка проверена и приложена к пакету.")
            elif action == "attach":
                attach_extract_to_package(self.object, user=request.user)
                messages.success(request, "Выкопировка приложена к пакету.")
            elif action == "recalc_mnp":
                snap = refresh_extract_mnp_intersections(self.object)
                messages.success(
                    request,
                    f"Пересечения с генпланом пересчитаны: зон {snap.get('count', 0)}"
                    f", жёстких {snap.get('hard_count', 0)}.",
                )
            else:
                messages.warning(request, "Неизвестное действие.")
        except InvestExtractError as exc:
            messages.error(request, str(exc))
        return redirect(reverse("invest-extract-detail", args=[self.object.pk]))


class InvestSiteExtractRequestView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        site = get_object_or_404(sites_for_membership(self.get_membership()), pk=kwargs["pk"])
        extract = ensure_extract_for_site(site, reason="manual_site", user=request.user, force=True)
        messages.success(request, "Запрос выкопировки создан.")
        if extract:
            return redirect(reverse("invest-extract-detail", args=[extract.pk]))
        return redirect(reverse("invest-site-detail", args=[site.pk]))


class InvestFgistpListView(InvestSubsystemMixin, ModulePermissionMixin, ListView):
    model = InvestFgistpRecord
    template_name = "invest/fgistp_list.html"
    context_object_name = "records"
    page_title = "Сведения ФГИС ТП"
    paginate_by = 50

    def get_queryset(self):
        membership = self.get_membership()
        qs = (
            InvestFgistpRecord.objects.filter(
                subsystem=membership.subsystem, site__in=sites_for_membership(membership)
            )
            .select_related("site", "site__organization", "project", "requested_by")
            .order_by("-updated_at")
        )
        status = (self.request.GET.get("status") or "").strip()
        if status:
            qs = qs.filter(status=status)
        org = (self.request.GET.get("org") or "").strip()
        if org.isdigit():
            qs = qs.filter(site__organization_id=int(org))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        ctx["status_choices"] = InvestFgistpRecord.Status.choices
        ctx["filter_status"] = self.request.GET.get("status") or ""
        ctx["filter_org"] = self.request.GET.get("org") or ""
        from delayu.models import Organization

        ctx["orgs"] = Organization.objects.filter(subsystem=membership.subsystem).order_by("name")
        ctx["can_change"] = user_can(self.request.user, self.module_code, "change")
        return ctx


class InvestFgistpDetailView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, DetailView):
    model = InvestFgistpRecord
    template_name = "invest/fgistp_detail.html"
    context_object_name = "record"
    page_title = "Сведения ФГИС ТП"
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        membership = self.get_membership()
        return InvestFgistpRecord.objects.filter(
            subsystem=membership.subsystem,
            site__in=sites_for_membership(membership),
        ).select_related("site", "site__organization", "project")

    def get(self, request, *args, **kwargs):
        """If record id missing but catalog document exists — open document card (demo UX)."""
        try:
            self.object = self.get_object()
        except Http404:
            membership = self.get_membership()
            document = (
                InvestFgistpDocument.objects.filter(
                    Q(subsystem=membership.subsystem) | Q(subsystem__isnull=True),
                    is_active=True,
                    pk=kwargs.get("pk"),
                ).first()
            )
            if document:
                messages.info(
                    request,
                    "Открыт документ демо-каталога ФГИС ТП (это не карточка привязанных сведений).",
                )
                return redirect(reverse("invest-fgistp-document-detail", args=[document.pk]))
            raise
        context = self.get_context_data(object=self.object)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["can_change"] = user_can(self.request.user, self.module_code, "change")
        geom = fgistp_geometry_for_map(self.object if self.object.geometry else None)
        ctx["fgistp_map"] = geom
        ctx["fgistp_map_json"] = json.dumps(geom, cls=DjangoJSONEncoder, ensure_ascii=False) if geom else "null"
        return ctx

    def post(self, request, *args, **kwargs):
        if not user_can(request.user, self.module_code, "change"):
            return _forbidden_with_message(request, "Недостаточно прав")
        self.object = self.get_object()
        action = (request.POST.get("action") or "").strip()
        try:
            if action == "request":
                ensure_fgistp_for_site(
                    self.object.site,
                    reason="manual",
                    user=request.user,
                    project=self.object.project,
                    force=True,
                )
                messages.success(request, "Запрос сведений ФГИС ТП создан/обновлён.")
            elif action == "upload":
                uploaded = request.FILES.get("file")
                if not uploaded:
                    messages.error(request, "Выберите файл.")
                else:
                    mark_fgistp_received(self.object, user=request.user, uploaded_file=uploaded)
                    messages.success(request, "Файл загружен, статус: получены.")
            elif action == "mock":
                generate_mock_zones(self.object, user=request.user)
                messages.success(request, "Mock-зоны ФГИС ТП сгенерированы.")
            elif action == "import_geometry":
                uploaded = request.FILES.get("geometry_file")
                raw = request.POST.get("geometry_json") or ""
                if uploaded:
                    import_fgistp_geometry(
                        self.object,
                        raw=uploaded.read(),
                        filename=uploaded.name,
                        user=request.user,
                    )
                elif raw.strip():
                    import_fgistp_geometry(self.object, raw=raw, filename="inline.geojson", user=request.user)
                else:
                    raise InvestFgistpError("Укажите GeoJSON или файл")
                messages.success(request, "Геометрия импортирована.")
            elif action == "verify":
                verify_fgistp(self.object, user=request.user, attach=True)
                messages.success(request, "Сведения проверены и приложены к пакету (ИСОГД).")
            elif action == "attach":
                attach_fgistp_to_package(self.object, user=request.user)
                messages.success(request, "Сведения приложены к пакету.")
            else:
                messages.warning(request, "Неизвестное действие.")
        except InvestFgistpError as exc:
            messages.error(request, str(exc))
        return redirect(reverse("invest-fgistp-detail", args=[self.object.pk]))


class InvestSiteFgistpRequestView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        site = get_object_or_404(sites_for_membership(self.get_membership()), pk=kwargs["pk"])
        record = ensure_fgistp_for_site(site, reason="manual_site", user=request.user, force=True)
        messages.success(request, "Запрос сведений ФГИС ТП создан.")
        if record:
            return redirect(reverse("invest-fgistp-detail", args=[record.pk]))
        return redirect(reverse("invest-site-detail", args=[site.pk]))


class InvestFgistpSearchView(InvestSubsystemMixin, ModulePermissionMixin, TemplateView):
    template_name = "invest/fgistp_search.html"
    page_title = "Поиск документов ФГИС ТП"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        q = (self.request.GET.get("q") or "").strip()
        level = (self.request.GET.get("level") or "").strip()
        site_id = (self.request.GET.get("site") or "").strip()
        results = []
        if q or self.request.GET.get("show_all"):
            results = search_fgistp_documents(subsystem=membership.subsystem, q=q, level=level)
        ctx["q"] = q
        ctx["filter_level"] = level
        ctx["level_choices"] = InvestFgistpDocument.Level.choices
        ctx["results"] = results
        ctx["can_change"] = user_can(self.request.user, self.module_code, "change")
        ctx["sites"] = sites_for_membership(membership).select_related("organization").order_by("cadastral_number")
        ctx["preselected_site"] = None
        if site_id.isdigit():
            ctx["preselected_site"] = sites_for_membership(membership).filter(pk=int(site_id)).first()
        ctx["demo_catalog_notice"] = True
        return ctx


class InvestFgistpDocumentDetailView(InvestSubsystemMixin, ModulePermissionMixin, DetailView):
    model = InvestFgistpDocument
    template_name = "invest/fgistp_document_detail.html"
    context_object_name = "document"
    page_title = "Документ ФГИС ТП"

    def get_queryset(self):
        membership = self.get_membership()
        return InvestFgistpDocument.objects.filter(
            Q(subsystem=membership.subsystem) | Q(subsystem__isnull=True),
            is_active=True,
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        membership = self.get_membership()
        ctx["can_change"] = user_can(self.request.user, self.module_code, "change")
        ctx["sites"] = sites_for_membership(membership).select_related("organization").order_by("cadastral_number")
        site_id = (self.request.GET.get("site") or "").strip()
        ctx["preselected_site"] = None
        if site_id.isdigit():
            ctx["preselected_site"] = sites_for_membership(membership).filter(pk=int(site_id)).first()
        geom = None
        if self.object.geometry:
            from delayu.services.invest_extracts import geojson_ring_to_yandex_coords

            coords = geojson_ring_to_yandex_coords(self.object.geometry)
            if len(coords) >= 3:
                geom = {"coords": coords, "name": self.object.title}
        ctx["doc_map"] = geom
        ctx["doc_map_json"] = json.dumps(geom, cls=DjangoJSONEncoder, ensure_ascii=False) if geom else "null"
        return ctx


class InvestFgistpDocumentAttachView(InvestForbiddenResponseMixin, InvestSubsystemMixin, ModulePermissionMixin, View):
    required_action = "change"

    def post(self, request, *args, **kwargs):
        membership = self.get_membership()
        document = get_object_or_404(
            InvestFgistpDocument.objects.filter(
                Q(subsystem=membership.subsystem) | Q(subsystem__isnull=True),
                is_active=True,
            ),
            pk=kwargs["pk"],
        )
        site_id = (request.POST.get("site_id") or "").strip()
        if not site_id.isdigit():
            messages.error(request, "Выберите площадку для привязки.")
            return redirect(reverse("invest-fgistp-document-detail", args=[document.pk]))
        site = get_object_or_404(sites_for_membership(membership), pk=int(site_id))
        try:
            record = attach_fgistp_document(document=document, site=site, user=request.user)
        except InvestFgistpError as exc:
            messages.error(request, str(exc))
            return redirect(f"{reverse('invest-fgistp-search')}?q={site.cadastral_number}&site={site.pk}")
        messages.success(request, f"Документ привязан к площадке {site.cadastral_number}.")
        return redirect(reverse("invest-fgistp-detail", args=[record.pk]))
