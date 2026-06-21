"""Экспорт отчётов УЖВ в XLSX, PDF и HTML."""
from __future__ import annotations

import csv
import io
import zipfile
from datetime import date
from xml.sax.saxutils import escape

from django.http import HttpResponse
from django.utils import timezone

from delayu.services.uzhv_reports import REPORT_BUILDERS, _csv_response_rows
from delayu.services.uzhv_report_forms import FORM_REPORT_CODES, build_form_report_rows


def csv_string_to_rows(content: str) -> list[list]:
    text = content.lstrip("\ufeff")
    return list(csv.reader(io.StringIO(text), delimiter=";"))


def build_report_rows(
    code: str,
    subsystem,
    *,
    period_start: date | None = None,
    period_end: date | None = None,
) -> tuple[str, list[list]]:
    if code in FORM_REPORT_CODES:
        return build_form_report_rows(code, subsystem, period_start, period_end)
    meta = REPORT_BUILDERS.get(code)
    if not meta:
        raise KeyError(code)
    title, builder, needs_period = meta
    if needs_period:
        csv_content = builder(subsystem, period_start, period_end)
    else:
        csv_content = builder(subsystem)
    return title, csv_string_to_rows(csv_content)


def _col_letter(index: int) -> str:
    result = ""
    n = index
    while True:
        result = chr(n % 26 + ord("A")) + result
        n = n // 26 - 1
        if n < 0:
            break
    return result


def rows_to_xlsx_bytes(rows: list[list], sheet_title: str = "Отчёт") -> bytes:
    """XLSX: openpyxl при наличии, иначе минимальный OOXML через stdlib."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]
        bold = Font(bold=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = bold
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        pass
    return _rows_to_xlsx_stdlib(rows, sheet_title)


def _rows_to_xlsx_stdlib(rows: list[list], sheet_title: str = "Отчёт") -> bytes:
    """XLSX без openpyxl — минимальный OOXML через stdlib."""
    sheet_rows = []
    for r_idx, row in enumerate(rows, 1):
        cells = []
        for c_idx, val in enumerate(row):
            ref = f"{_col_letter(c_idx)}{r_idx}"
            text = escape(str(val) if val is not None else "")
            cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{text}</t></is></c>')
        sheet_rows.append(f'<row r="{r_idx}">{"".join(cells)}</row>')

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(sheet_rows)}</sheetData></worksheet>"
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets><sheet name="{escape(sheet_title[:31])}" sheetId="1" r:id="rId1"/></sheets>'
        "</workbook>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )
    rels_root = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    rels_wb = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        "</Relationships>"
    )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", rels_root)
        z.writestr("xl/workbook.xml", workbook_xml)
        z.writestr("xl/_rels/workbook.xml.rels", rels_wb)
        z.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return buf.getvalue()


def _find_header_row_index(rows: list[list], marker: str = "№ п/п") -> int:
    for i, row in enumerate(rows):
        if row and str(row[0]).strip().startswith(marker):
            return i
    return 0


def rows_to_formatted_xlsx_bytes(
    rows: list[list],
    sheet_title: str = "Отчёт",
    *,
    header_row_index: int | None = None,
) -> bytes:
    """XLSX с шапкой отчёта и оформленной строкой заголовков таблицы."""
    if header_row_index is None:
        header_row_index = _find_header_row_index(rows)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]
        bold = Font(bold=True)
        title_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        max_col = max((len(r) for r in rows if r), default=1)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1 and c_idx == 1:
                    cell.font = title_font
                if r_idx - 1 == header_row_index:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                elif r_idx - 1 > header_row_index and row and str(row[0]).startswith(("Итого", "Сводка")):
                    cell.font = bold

        if max_col > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.freeze_panes = ws.cell(row=header_row_index + 2, column=1)
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 16 if col > 1 else 6

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        pass
    return rows_to_xlsx_bytes(rows, sheet_title)


def _pdf_font_name() -> str:
    """Шрифт с поддержкой кириллицы (Windows Arial) или Helvetica."""
    import os

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    for path in (
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ):
        if os.path.isfile(path):
            try:
                pdfmetrics.registerFont(TTFont("UzHVExport", path))
                pdfmetrics.registerFont(TTFont("UzHVExport-Bold", path))
                return "UzHVExport"
            except Exception:
                continue
    return "Helvetica"


def rows_to_pdf_bytes(title: str, rows: list[list]) -> bytes:
    """PDF через reportlab; при отсутствии пакета — HTML для печати в PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        font = _pdf_font_name()
        bold_font = "UzHVExport-Bold" if font == "UzHVExport" else "Helvetica-Bold"

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=24, rightMargin=24)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "UzHVTitle", parent=styles["Title"], fontName=font
        )
        normal_style = ParagraphStyle(
            "UzHVNormal", parent=styles["Normal"], fontName=font
        )
        story = [
            Paragraph(title, title_style),
            Spacer(1, 8),
            Paragraph(
                f"Сформировано: {timezone.now():%d.%m.%Y %H:%M}",
                normal_style,
            ),
            Spacer(1, 12),
        ]
        if rows:
            table = Table(rows, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e88e5")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), bold_font),
                        ("FONTNAME", (0, 1), (-1, -1), font),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(table)
        doc.build(story)
        return buf.getvalue()
    except ImportError:
        return rows_to_printable_html_bytes(title, rows)


def rows_to_printable_html_bytes(title: str, rows: list[list]) -> bytes:
    """HTML-версия отчёта (UTF-8) для печати в PDF из браузера."""
    head = rows[0] if rows else []
    body_rows = rows[1:] if len(rows) > 1 else []
    th = "".join(f"<th>{escape(str(c))}</th>" for c in head)
    trs = []
    for row in body_rows:
        tds = "".join(f"<td>{escape(str(c))}</td>" for c in row)
        trs.append(f"<tr>{tds}</tr>")
    html = f"""<!DOCTYPE html>
<html lang="ru"><head><meta charset="utf-8"><title>{escape(title)}</title>
<style>
body {{ font-family: Arial, sans-serif; margin: 24px; }}
h1 {{ font-size: 18px; }}
table {{ border-collapse: collapse; width: 100%; font-size: 12px; margin-top: 16px; }}
th, td {{ border: 1px solid #ccc; padding: 6px 8px; text-align: left; vertical-align: top; }}
th {{ background: #1e88e5; color: #fff; }}
.meta {{ color: #666; font-size: 12px; }}
@media print {{ body {{ margin: 0; }} }}
</style></head><body>
<h1>{escape(title)}</h1>
<p class="meta">Сформировано: {timezone.now():%d.%m.%Y %H:%M} · АИС УЖВ / ДелаЮ</p>
<table><thead><tr>{th}</tr></thead><tbody>{''.join(trs)}</tbody></table>
</body></html>"""
    return html.encode("utf-8")


def http_export_report(
    *,
    code: str,
    subsystem,
    fmt: str,
    period_start: date | None = None,
    period_end: date | None = None,
) -> HttpResponse:
    title, rows = build_report_rows(
        code, subsystem, period_start=period_start, period_end=period_end
    )
    stamp = timezone.now().strftime("%Y%m%d")
    safe_code = code.replace("/", "-")

    if fmt == "xlsx":
        if code in FORM_REPORT_CODES:
            content = rows_to_formatted_xlsx_bytes(rows, sheet_title=code)
        else:
            content = rows_to_xlsx_bytes(rows, sheet_title=code)
        resp = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="uzhv-{safe_code}-{stamp}.xlsx"'
        return resp

    if fmt == "pdf":
        content = rows_to_pdf_bytes(title, rows)
        if content[:5] == b"<!DOC":
            resp = HttpResponse(content, content_type="text/html; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="uzhv-{safe_code}-{stamp}.html"'
            return resp
        resp = HttpResponse(content, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="uzhv-{safe_code}-{stamp}.pdf"'
        return resp

    csv_content = _csv_response_rows(rows)
    resp = HttpResponse(csv_content, content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="uzhv-{safe_code}-{stamp}.csv"'
    return resp
